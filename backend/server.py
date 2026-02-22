from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
from uuid import uuid4
from datetime import datetime, timezone
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from passlib.context import CryptContext
from jose import JWTError, jwt
from datetime import timedelta

from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Import WhatsApp routes
from whatsapp_routes import whatsapp_router


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# ========== MODELS ==========

# Product Models
class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    name: str
    hsn_code: Optional[str] = ""
    unit: str = "pcs"
    price: float
    gst_rate: float = 0.0  # 0, 5, 12, 18, 28
    stock_quantity: float = 0.0
    min_stock_level: float = 0.0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductCreate(BaseModel):
    name: str
    hsn_code: Optional[str] = ""
    unit: str = "pcs"
    price: float
    gst_rate: float = 0.0
    stock_quantity: float = 0.0
    min_stock_level: float = 0.0

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    hsn_code: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    gst_rate: Optional[float] = None
    stock_quantity: Optional[float] = None
    min_stock_level: Optional[float] = None


# Customer Models
class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    name: str
    address: Optional[str] = ""
    phone: Optional[str] = ""
    gst_number: Optional[str] = ""
    email: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomerCreate(BaseModel):
    name: str
    address: Optional[str] = ""
    phone: Optional[str] = ""
    gst_number: Optional[str] = ""
    email: Optional[str] = ""

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    gst_number: Optional[str] = None
    email: Optional[str] = None


# Supplier Models
class Supplier(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    name: str
    address: Optional[str] = ""
    phone: Optional[str] = ""
    gst_number: Optional[str] = ""
    email: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SupplierCreate(BaseModel):
    name: str
    address: Optional[str] = ""
    phone: Optional[str] = ""
    gst_number: Optional[str] = ""
    email: Optional[str] = ""

class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    gst_number: Optional[str] = None
    email: Optional[str] = None


# Raw Material Models
class RawMaterial(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    name: str
    hsn_code: Optional[str] = ""
    unit: str = "kg"
    purchase_price: float
    gst_rate: float = 0.0
    stock_quantity: float = 0.0
    min_stock_level: float = 0.0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RawMaterialCreate(BaseModel):
    name: str
    hsn_code: Optional[str] = ""
    unit: str = "kg"
    purchase_price: float = 0.0
    gst_rate: float = 0.0
    stock_quantity: float = 0.0
    min_stock_level: float = 0.0

class RawMaterialUpdate(BaseModel):
    name: Optional[str] = None
    hsn_code: Optional[str] = None
    unit: Optional[str] = None
    purchase_price: Optional[float] = None
    gst_rate: Optional[float] = None
    stock_quantity: Optional[float] = None
    min_stock_level: Optional[float] = None


# Packing Material Models
class PackingMaterial(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    name: str
    unit: str = "pcs"
    purchase_price: float
    gst_rate: float = 0.0
    stock_quantity: float = 0.0
    min_stock_level: float = 0.0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PackingMaterialCreate(BaseModel):
    name: str
    unit: str = "pcs"
    purchase_price: float = 0.0
    gst_rate: float = 0.0
    stock_quantity: float = 0.0
    min_stock_level: float = 0.0

class PackingMaterialUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    purchase_price: Optional[float] = None
    gst_rate: Optional[float] = None
    stock_quantity: Optional[float] = None
    min_stock_level: Optional[float] = None


# Company Settings Models
class CompanySettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    company_name: str = "My Company"
    address: Optional[str] = ""
    city: Optional[str] = ""
    state: Optional[str] = ""
    pincode: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    website: Optional[str] = ""
    gst_number: Optional[str] = ""
    pan_number: Optional[str] = ""
    logo_url: Optional[str] = ""
    bank_name: Optional[str] = ""
    account_number: Optional[str] = ""
    account_holder: Optional[str] = ""
    ifsc_code: Optional[str] = ""
    branch: Optional[str] = ""
    terms_and_conditions: Optional[str] = ""
    invoice_footer: Optional[str] = "Thank you for your business!"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CompanySettingsCreate(BaseModel):
    company_name: str = "My Company"
    address: Optional[str] = ""
    city: Optional[str] = ""
    state: Optional[str] = ""
    pincode: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    website: Optional[str] = ""
    gst_number: Optional[str] = ""
    pan_number: Optional[str] = ""
    logo_url: Optional[str] = ""
    bank_name: Optional[str] = ""
    account_number: Optional[str] = ""
    account_holder: Optional[str] = ""
    ifsc_code: Optional[str] = ""
    branch: Optional[str] = ""
    terms_and_conditions: Optional[str] = ""
    invoice_footer: Optional[str] = "Thank you for your business!"


# Invoice Models
class InvoiceItem(BaseModel):
    product_id: str
    product_name: str
    hsn_code: Optional[str] = ""
    quantity: float
    unit: str
    price: float
    discount_percent: float = 0.0
    gst_rate: float = 0.0

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    invoice_number: str
    invoice_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    invoice_status: str = "draft"  # draft, confirmed, dispatched, delivered
    customer_id: str
    customer_name: str
    customer_address: Optional[str] = ""
    customer_phone: Optional[str] = ""
    customer_gst: Optional[str] = ""
    buyer_order_no: Optional[str] = ""
    vehicle_no: Optional[str] = ""
    payment_terms: Optional[str] = "Immediate"
    items: List[InvoiceItem]
    subtotal: float
    total_discount: float
    overall_discount_type: Optional[str] = "percentage"  # percentage, amount
    overall_discount_value: float = 0.0
    overall_discount_amount: float = 0.0
    taxable_amount: float
    is_interstate: bool = False
    cgst_amount: float = 0.0
    sgst_amount: float = 0.0
    igst_amount: float = 0.0
    total_gst: float
    grand_total: float
    payment_status: str = "unpaid"  # unpaid, partial, paid
    stock_updated: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvoiceCreate(BaseModel):
    customer_id: str
    customer_name: str
    customer_address: Optional[str] = ""
    customer_phone: Optional[str] = ""
    customer_gst: Optional[str] = ""
    buyer_order_no: Optional[str] = ""
    vehicle_no: Optional[str] = ""
    payment_terms: Optional[str] = "Immediate"
    invoice_status: str = "draft"
    items: List[InvoiceItem]
    overall_discount_type: Optional[str] = "percentage"
    overall_discount_value: float = 0.0
    payment_status: str = "unpaid"


# Quotation Models (similar to Invoice but for quotes)
class Quotation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    quotation_number: str
    quotation_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    valid_until: Optional[datetime] = None
    quotation_status: str = "draft"  # draft, sent, accepted, rejected, converted
    customer_id: str
    customer_name: str
    customer_address: Optional[str] = ""
    customer_phone: Optional[str] = ""
    customer_gst: Optional[str] = ""
    items: List[InvoiceItem]
    subtotal: float
    total_discount: float
    overall_discount_type: Optional[str] = "percentage"
    overall_discount_value: float = 0.0
    overall_discount_amount: float = 0.0
    taxable_amount: float
    is_interstate: bool = False
    cgst_amount: float = 0.0
    sgst_amount: float = 0.0
    igst_amount: float = 0.0
    total_gst: float
    grand_total: float
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuotationCreate(BaseModel):
    customer_id: str
    customer_name: str
    customer_address: Optional[str] = ""
    customer_phone: Optional[str] = ""
    customer_gst: Optional[str] = ""
    valid_until: Optional[datetime] = None
    items: List[InvoiceItem]
    overall_discount_type: Optional[str] = "percentage"
    overall_discount_value: float = 0.0
    notes: Optional[str] = ""


# Sales Order Models
class SalesOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    so_number: str
    so_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    so_status: str = "draft"  # draft, pending_approval, approved, rejected, invoiced, cancelled
    approval_status: str = "pending"  # pending, approved, rejected
    approved_by: Optional[str] = None  # Inventory employee name/ID
    approved_at: Optional[datetime] = None
    rejection_reason: Optional[str] = None
    production_status: str = "pending"  # pending, in_progress, completed
    dispatch_status: str = "pending"  # pending, ready, dispatched, delivered
    customer_id: str
    customer_name: str
    customer_address: Optional[str] = ""
    customer_phone: Optional[str] = ""
    customer_gst: Optional[str] = ""
    buyer_order_no: Optional[str] = ""
    expected_delivery_date: Optional[datetime] = None
    items: List[InvoiceItem]
    subtotal: float
    total_discount: float
    overall_discount_type: Optional[str] = "percentage"
    overall_discount_value: float = 0.0
    overall_discount_amount: float = 0.0
    taxable_amount: float
    is_interstate: bool = False
    cgst_amount: float = 0.0
    sgst_amount: float = 0.0
    igst_amount: float = 0.0
    total_gst: float
    grand_total: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SalesOrderCreate(BaseModel):
    customer_id: str
    customer_name: str
    customer_address: Optional[str] = ""
    customer_phone: Optional[str] = ""
    customer_gst: Optional[str] = ""
    buyer_order_no: Optional[str] = ""
    expected_delivery_date: Optional[datetime] = None
    items: List[InvoiceItem]
    overall_discount_type: Optional[str] = "percentage"
    overall_discount_value: float = 0.0


# Delivery Challan Models
class DeliveryChallan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    challan_number: str
    challan_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    sales_order_id: Optional[str] = ""
    sales_order_number: Optional[str] = ""
    invoice_id: Optional[str] = ""
    invoice_number: Optional[str] = ""
    customer_id: str
    customer_name: str
    customer_address: Optional[str] = ""
    vehicle_no: Optional[str] = ""
    transporter_name: Optional[str] = ""
    dispatch_status: str = "pending"  # pending, dispatched, in_transit, delivered
    items: List[InvoiceItem]
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DeliveryChallanCreate(BaseModel):
    sales_order_id: Optional[str] = ""
    sales_order_number: Optional[str] = ""
    invoice_id: Optional[str] = ""
    invoice_number: Optional[str] = ""
    customer_id: str
    customer_name: str
    customer_address: Optional[str] = ""
    vehicle_no: Optional[str] = ""
    transporter_name: Optional[str] = ""
    items: List[InvoiceItem]


# Credit Note Models (for returns)
class CreditNote(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    credit_note_number: str
    credit_note_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    invoice_id: str
    invoice_number: str
    customer_id: str
    customer_name: str
    reason: Optional[str] = ""
    items: List[InvoiceItem]
    subtotal: float
    total_discount: float
    taxable_amount: float
    is_interstate: bool = False
    cgst_amount: float = 0.0
    sgst_amount: float = 0.0
    igst_amount: float = 0.0
    total_gst: float
    credit_amount: float
    stock_restored: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CreditNoteCreate(BaseModel):
    invoice_id: str
    invoice_number: str
    customer_id: str
    customer_name: str
    reason: Optional[str] = ""
    items: List[InvoiceItem]


# Payment Models
class PaymentAllocation(BaseModel):
    invoice_id: str
    invoice_number: str
    invoice_type: str  # sales_invoice, purchase_invoice
    allocated_amount: float

class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    payment_number: str
    payment_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    payment_type: str  # receive (from customer), pay (to supplier)
    partner_id: str
    partner_name: str
    partner_type: str  # customer, supplier
    payment_method: str  # cash, bank, cheque, upi, online
    payment_amount: float
    bank_reference: Optional[str] = ""
    cheque_number: Optional[str] = ""
    upi_transaction_id: Optional[str] = ""
    allocations: List[PaymentAllocation] = []
    unallocated_amount: float = 0.0
    memo: Optional[str] = ""
    status: str = "posted"  # draft, posted, reconciled
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentCreate(BaseModel):
    payment_type: str
    partner_id: str
    partner_name: str
    partner_type: str
    payment_method: str
    payment_amount: float
    bank_reference: Optional[str] = ""
    cheque_number: Optional[str] = ""
    upi_transaction_id: Optional[str] = ""
    allocations: List[PaymentAllocation] = []
    memo: Optional[str] = ""
    status: str = "posted"


# Journal Entry Models
class JournalEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    entry_number: str
    entry_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    entry_type: str  # freight, discount, other_charges, adjustment
    customer_id: str
    customer_name: str
    description: str
    amount: float  # Positive for charges (debit), Negative for discounts/credits (credit)
    reference_type: Optional[str] = ""  # invoice, sales_order, delivery_challan
    reference_id: Optional[str] = ""
    reference_number: Optional[str] = ""
    status: str = "posted"  # draft, posted
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class JournalEntryCreate(BaseModel):
    entry_type: str
    customer_id: str
    customer_name: str
    description: str
    amount: float
    reference_type: Optional[str] = ""
    reference_id: Optional[str] = ""
    reference_number: Optional[str] = ""
    status: str = "posted"


# Purchase Order Models
class PurchaseOrderItem(BaseModel):
    item_id: str
    item_name: str
    item_type: str  # raw_material, packing_material
    hsn_code: Optional[str] = ""
    quantity: float
    unit: str
    price: float
    gst_rate: float = 0.0

class PurchaseOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    po_number: str
    po_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    supplier_id: str
    supplier_name: str
    supplier_address: Optional[str] = ""
    supplier_phone: Optional[str] = ""
    supplier_gst: Optional[str] = ""
    items: List[PurchaseOrderItem]
    subtotal: float
    total_gst: float
    grand_total: float
    status: str = "draft"  # draft, confirmed, received, cancelled
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PurchaseOrderCreate(BaseModel):
    supplier_id: str
    supplier_name: str
    supplier_address: Optional[str] = ""
    supplier_phone: Optional[str] = ""
    supplier_gst: Optional[str] = ""
    items: List[PurchaseOrderItem]
    status: str = "draft"


# Purchase Invoice Models
class PurchaseInvoiceItem(BaseModel):
    item_id: str
    item_name: str
    item_type: str  # raw_material, packing_material
    hsn_code: Optional[str] = ""
    quantity: float
    unit: str
    price: float
    discount_percent: float = 0.0
    gst_rate: float = 0.0

class PurchaseInvoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    invoice_number: str
    invoice_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    supplier_id: str
    supplier_name: str
    supplier_address: Optional[str] = ""
    supplier_phone: Optional[str] = ""
    supplier_gst: Optional[str] = ""
    supplier_invoice_no: Optional[str] = ""
    items: List[PurchaseInvoiceItem]
    subtotal: float
    total_discount: float
    taxable_amount: float
    cgst_amount: float
    sgst_amount: float
    total_gst: float
    grand_total: float
    payment_status: str = "unpaid"
    stock_updated: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PurchaseInvoiceCreate(BaseModel):
    supplier_id: str
    supplier_name: str
    supplier_address: Optional[str] = ""
    supplier_phone: Optional[str] = ""
    supplier_gst: Optional[str] = ""
    supplier_invoice_no: Optional[str] = ""
    items: List[PurchaseInvoiceItem]
    payment_status: str = "unpaid"




# BOM (Bill of Materials) Models
class BOMMaterial(BaseModel):
    material_type: str  # 'raw' or 'packing'
    material_id: str
    material_name: str
    quantity: float
    unit: str

class BOM(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    product_id: str
    product_name: str
    materials: List[BOMMaterial]
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BOMCreate(BaseModel):
    product_id: str
    product_name: str
    materials: List[BOMMaterial]
    notes: Optional[str] = ""

class BOMUpdate(BaseModel):
    materials: Optional[List[BOMMaterial]] = None
    notes: Optional[str] = None


# Production Order Models
class ProductionOrderMaterial(BaseModel):
    material_type: str  # 'raw' or 'packing'
    material_id: str
    material_name: str
    required_quantity: float
    allocated_quantity: float = 0.0
    unit: str

class ProductionOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    order_number: str
    product_id: str
    product_name: str
    quantity_to_produce: float
    bom_id: Optional[str] = None
    status: str = "draft"  # draft, scheduled, in_progress, completed, cancelled
    materials_required: List[ProductionOrderMaterial] = []
    start_date: Optional[datetime] = None
    completion_date: Optional[datetime] = None
    notes: Optional[str] = ""
    sales_order_id: Optional[str] = None  # Link to sales order
    sales_order_number: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductionOrderCreate(BaseModel):
    product_id: str
    product_name: str
    quantity_to_produce: float
    bom_id: Optional[str] = None


# Expense Models
class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    expense_number: str
    expense_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    category: str  # office_supplies, utilities, rent, salaries, travel, marketing, etc.
    description: str
    amount: float
    payment_method: str = "cash"  # cash, bank, upi, card
    vendor_name: Optional[str] = ""
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExpenseCreate(BaseModel):
    category: str
    description: str
    amount: float
    payment_method: str = "cash"
    vendor_name: Optional[str] = ""
    notes: Optional[str] = ""


# Day Book Models (Manual Entry)
class DayBookEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    date: datetime
    description: str
    purpose: Optional[str] = ""
    debit: float = 0.0  # Money In
    credit: float = 0.0  # Money Out
    balance: float = 0.0  # Running balance
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DayBookEntryCreate(BaseModel):
    date: str
    description: str
    purpose: Optional[str] = ""
    debit: float = 0.0
    credit: float = 0.0




    notes: Optional[str] = ""

class ProductionOrderUpdate(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None



class MaterialRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    request_number: str
    production_order_id: str
    production_order_number: str
    requested_by: str
    request_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "pending"  # pending, approved, rejected
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    rejection_reason: Optional[str] = None
    materials: List[dict]  # List of {material_type, material_id, material_name, quantity_requested}
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MaterialRequestCreate(BaseModel):
    production_order_id: str
    requested_by: str


class PurchaseRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    request_number: str
    requested_by: str  # Inventory Manager name
    request_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "pending"  # pending, approved, rejected
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    rejection_reason: Optional[str] = None
    items: List[dict]  # List of {material_type, material_id, material_name, quantity, estimated_cost}
    total_estimated_cost: float = 0.0
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PurchaseRequestCreate(BaseModel):
    requested_by: str
    items: List[dict]
    total_estimated_cost: float = 0.0
    notes: Optional[str] = ""


# Stock Inward Models
class StockInward(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    material_type: str  # raw_material or packing_material
    material_id: str
    material_name: str
    quantity_added: float
    unit: str
    added_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StockInwardCreate(BaseModel):
    material_type: str
    material_id: str
    quantity_added: float


# Supplier Price Models
class SupplierPrice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    supplier_id: str
    supplier_name: str
    material_type: str  # raw_material or packing_material
    material_id: str
    material_name: str
    unit: str
    price: float
    lead_time_days: Optional[int] = None  # Delivery time
    minimum_order_qty: Optional[float] = None
    notes: Optional[str] = ""
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SupplierPriceCreate(BaseModel):
    supplier_id: str
    material_type: str
    material_id: str
    price: float
    lead_time_days: Optional[int] = None
    minimum_order_qty: Optional[float] = None
    notes: Optional[str] = ""

class SupplierPriceUpdate(BaseModel):
    price: Optional[float] = None
    lead_time_days: Optional[int] = None
    minimum_order_qty: Optional[float] = None
    notes: Optional[str] = None


# Financial Transaction Models
class FinancialTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    transaction_type: str  # payment_received, payment_made, purchase, sale
    category: str  # sales, purchase, expense, investment
    amount: float
    description: str
    reference_id: Optional[str] = None  # Link to invoice, PO, etc.
    reference_type: Optional[str] = None  # invoice, purchase_order, expense
    party_name: Optional[str] = ""  # Customer or Supplier name
    payment_method: Optional[str] = "cash"  # cash, bank, upi, cheque
    transaction_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FinancialTransactionCreate(BaseModel):
    transaction_type: str
    category: str
    amount: float
    description: str
    reference_id: Optional[str] = None
    reference_type: Optional[str] = None
    party_name: Optional[str] = ""
    payment_method: Optional[str] = "cash"
    transaction_date: Optional[datetime] = None
    notes: Optional[str] = ""



# User and Authentication Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    username: str
    email: Optional[str] = ""
    hashed_password: str
    full_name: str
    role: str  # admin, purchase_manager, production_manager, finance_manager, inventory_manager
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    password: str
    email: Optional[str] = ""
    full_name: str
    role: str

class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: dict

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


# ========== USER & ROLE MANAGEMENT MODELS ==========

class Permission(BaseModel):
    module: str  # inventory, manufacturing, sales, purchase, finance, hr
    pages: List[str]  # list of page paths user can access

class Role(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    name: str
    description: Optional[str] = ""
    permissions: List[Permission] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RoleCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    permissions: List[Permission] = []

class UserWithRole(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    username: str
    email: str
    full_name: str
    role: str  # admin, hr_manager, finance_manager, inventory_manager, etc.
    role_id: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    email: str
    full_name: str
    password: str
    role: str
    role_id: Optional[str] = None
    is_active: bool = True


# ========== HR MODULE MODELS ==========

# Department Models
class Department(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    name: str
    description: Optional[str] = ""
    manager_id: Optional[str] = None
    manager_name: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DepartmentCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    manager_id: Optional[str] = None
    manager_name: Optional[str] = ""


# Employee Models
class Employee(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    employee_code: str
    first_name: str
    last_name: str
    email: str
    phone: str
    date_of_birth: Optional[datetime] = None
    gender: Optional[str] = ""
    address: Optional[str] = ""
    department_id: str
    department_name: str
    designation: str
    date_of_joining: datetime
    employment_type: str = "full_time"  # full_time, part_time, contract, intern
    status: str = "active"  # active, inactive, on_leave, terminated
    reporting_manager_id: Optional[str] = None
    reporting_manager_name: Optional[str] = ""
    salary: float = 0.0
    bank_name: Optional[str] = ""
    account_number: Optional[str] = ""
    ifsc_code: Optional[str] = ""
    pan_number: Optional[str] = ""
    aadhar_number: Optional[str] = ""
    emergency_contact_name: Optional[str] = ""
    emergency_contact_number: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployeeCreate(BaseModel):
    first_name: str
    last_name: str
    email: str
    phone: str
    date_of_birth: Optional[datetime] = None
    gender: Optional[str] = ""
    address: Optional[str] = ""
    department_id: str
    department_name: str
    designation: str
    date_of_joining: datetime
    employment_type: str = "full_time"
    reporting_manager_id: Optional[str] = None
    reporting_manager_name: Optional[str] = ""
    salary: float = 0.0
    bank_name: Optional[str] = ""
    account_number: Optional[str] = ""
    ifsc_code: Optional[str] = ""
    pan_number: Optional[str] = ""
    aadhar_number: Optional[str] = ""
    emergency_contact_name: Optional[str] = ""
    emergency_contact_number: Optional[str] = ""


# Attendance Models
class Attendance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    employee_id: str
    employee_name: str
    date: datetime
    check_in: Optional[datetime] = None
    check_out: Optional[datetime] = None
    status: str = "present"  # present, absent, half_day, leave, holiday
    work_hours: float = 0.0
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AttendanceCreate(BaseModel):
    employee_id: str
    employee_name: str
    date: datetime
    check_in: Optional[datetime] = None
    check_out: Optional[datetime] = None
    status: str = "present"
    work_hours: float = 0.0
    notes: Optional[str] = ""


# Leave Type Models
class LeaveType(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    name: str
    description: Optional[str] = ""
    days_per_year: int
    is_paid: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeaveTypeCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    days_per_year: int
    is_paid: bool = True


# Leave Request Models
class LeaveRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    employee_id: str
    employee_name: str
    leave_type_id: str
    leave_type_name: str
    start_date: datetime
    end_date: datetime
    days_requested: float
    reason: str
    status: str = "pending"  # pending, approved, rejected
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    rejection_reason: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeaveRequestCreate(BaseModel):
    employee_id: str
    employee_name: str
    leave_type_id: str
    leave_type_name: str
    start_date: datetime
    end_date: datetime
    days_requested: float
    reason: str


# Payslip Models
class PayslipComponent(BaseModel):
    component_name: str
    component_type: str  # earning, deduction
    amount: float

class Payslip(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    employee_id: str
    employee_name: str
    month: int
    year: int
    basic_salary: float
    components: List[PayslipComponent] = []
    total_earnings: float
    total_deductions: float
    net_salary: float
    status: str = "draft"  # draft, processed, paid
    payment_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PayslipCreate(BaseModel):
    employee_id: str
    employee_name: str
    month: int
    year: int
    basic_salary: float
    components: List[PayslipComponent] = []


# Performance Review Models
class PerformanceReview(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    employee_id: str
    employee_name: str
    review_period_start: datetime
    review_period_end: datetime
    reviewer_id: str
    reviewer_name: str
    overall_rating: float = 0.0  # 1-5 scale
    technical_skills: float = 0.0
    communication: float = 0.0
    teamwork: float = 0.0
    punctuality: float = 0.0
    quality_of_work: float = 0.0
    strengths: Optional[str] = ""
    areas_of_improvement: Optional[str] = ""
    comments: Optional[str] = ""
    status: str = "draft"  # draft, submitted, acknowledged
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PerformanceReviewCreate(BaseModel):
    employee_id: str
    employee_name: str
    review_period_start: datetime
    review_period_end: datetime
    reviewer_id: str
    reviewer_name: str
    overall_rating: float = 0.0
    technical_skills: float = 0.0
    communication: float = 0.0
    teamwork: float = 0.0
    punctuality: float = 0.0
    quality_of_work: float = 0.0
    strengths: Optional[str] = ""
    areas_of_improvement: Optional[str] = ""
    comments: Optional[str] = ""


# Goal/KPI Models
class Goal(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    employee_id: str
    employee_name: str
    title: str
    description: str
    target_date: datetime
    status: str = "in_progress"  # not_started, in_progress, completed, cancelled
    progress_percentage: float = 0.0
    assigned_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class GoalCreate(BaseModel):
    employee_id: str
    employee_name: str
    title: str
    description: str
    target_date: datetime
    assigned_by: str


# Employee Document Models
class EmployeeDocument(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    employee_id: str
    employee_name: str
    document_type: str  # resume, id_proof, address_proof, certificates, offer_letter, contract
    document_name: str
    document_url: str
    uploaded_by: str
    uploaded_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployeeDocumentCreate(BaseModel):
    employee_id: str
    employee_name: str
    document_type: str
    document_name: str
    document_url: str
    uploaded_by: str


# Job Posting Models
class JobPosting(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    job_title: str
    department_id: str
    department_name: str
    job_description: str
    requirements: str
    experience_required: str
    salary_range: Optional[str] = ""
    employment_type: str = "full_time"
    location: str
    status: str = "active"  # active, closed, on_hold
    posted_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    closing_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class JobPostingCreate(BaseModel):
    job_title: str
    department_id: str
    department_name: str
    job_description: str
    requirements: str
    experience_required: str
    salary_range: Optional[str] = ""
    employment_type: str = "full_time"
    location: str
    closing_date: Optional[datetime] = None


# Candidate Models
class Candidate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    job_posting_id: str
    job_title: str
    first_name: str
    last_name: str
    email: str
    phone: str
    resume_url: Optional[str] = ""
    cover_letter: Optional[str] = ""
    current_ctc: Optional[float] = 0.0
    expected_ctc: Optional[float] = 0.0
    notice_period: Optional[str] = ""
    status: str = "applied"  # applied, shortlisted, interview_scheduled, interviewed, offered, rejected, hired
    applied_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CandidateCreate(BaseModel):
    job_posting_id: str
    job_title: str
    first_name: str
    last_name: str
    email: str
    phone: str
    resume_url: Optional[str] = ""
    cover_letter: Optional[str] = ""
    current_ctc: Optional[float] = 0.0
    expected_ctc: Optional[float] = 0.0
    notice_period: Optional[str] = ""


# Interview Models
class Interview(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    candidate_id: str
    candidate_name: str
    job_title: str
    interview_date: datetime
    interview_type: str  # phone, video, in_person, technical, hr
    interviewer_id: str
    interviewer_name: str
    location: Optional[str] = ""
    meeting_link: Optional[str] = ""
    status: str = "scheduled"  # scheduled, completed, cancelled, rescheduled
    feedback: Optional[str] = ""
    rating: Optional[float] = 0.0
    result: Optional[str] = ""  # selected, rejected, on_hold
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InterviewCreate(BaseModel):
    candidate_id: str
    candidate_name: str
    job_title: str
    interview_date: datetime
    interview_type: str
    interviewer_id: str
    interviewer_name: str
    location: Optional[str] = ""
    meeting_link: Optional[str] = ""

# JWT settings
SECRET_KEY = "your-secret-key-change-this-in-production-2024"  # Change this in production
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440  # 24 hours

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt



# ========== HELPER FUNCTIONS ==========

async def generate_invoice_number():
    """Generate invoice number in format INV-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    # Find the last invoice for today
    last_invoice = await db.invoices.find_one(
        {"invoice_number": {"$regex": f"^INV-{date_str}"}},
        sort=[("invoice_number", -1)]
    )
    
    if last_invoice:
        last_num = int(last_invoice["invoice_number"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"INV-{date_str}-{new_num:04d}"


async def generate_po_number():
    """Generate PO number in format PO-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    last_po = await db.purchase_orders.find_one(
        {"po_number": {"$regex": f"^PO-{date_str}"}},
        sort=[("po_number", -1)]
    )
    
    if last_po:
        last_num = int(last_po["po_number"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"PO-{date_str}-{new_num:04d}"


async def generate_purchase_invoice_number():
    """Generate purchase invoice number in format PINV-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    last_invoice = await db.purchase_invoices.find_one(
        {"invoice_number": {"$regex": f"^PINV-{date_str}"}},
        sort=[("invoice_number", -1)]
    )
    
    if last_invoice:
        last_num = int(last_invoice["invoice_number"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"PINV-{date_str}-{new_num:04d}"


async def update_stock_on_purchase(items):
    """Update stock when purchase invoice is created"""
    for item in items:
        if item.item_type == "raw_material":
            await db.raw_materials.update_one(
                {"id": item.item_id},
                {"$inc": {"stock_quantity": item.quantity}}
            )
        elif item.item_type == "packing_material":
            await db.packing_materials.update_one(
                {"id": item.item_id},
                {"$inc": {"stock_quantity": item.quantity}}
            )


async def update_stock_on_sale(items):
    """Deduct stock when sales invoice is created"""
    for item in items:
        # Check if enough stock available
        product = await db.products.find_one({"id": item.product_id})
        if not product:
            raise HTTPException(status_code=404, detail=f"Product {item.product_name} not found")
        
        if product.get('stock_quantity', 0) < item.quantity:
            raise HTTPException(
                status_code=400, 
                detail=f"Insufficient stock for {item.product_name}. Available: {product.get('stock_quantity', 0)}, Required: {item.quantity}"
            )
        
        # Deduct stock
        await db.products.update_one(
            {"id": item.product_id},
            {"$inc": {"stock_quantity": -item.quantity}}
        )


def detect_interstate(customer_gst: str, company_gst: str = None) -> bool:
    """Detect if transaction is interstate based on GST numbers
    First 2 digits of GST number represent state code"""
    if not customer_gst or len(customer_gst) < 2:
        return False
    
    # If company_gst not provided, use default (will be fetched from settings in real scenario)
    if not company_gst:
        company_gst = "29AAAAA1234A1Z5"  # Default Karnataka
    
    if len(company_gst) < 2:
        return False
    
    customer_state = customer_gst[:2]
    company_state = company_gst[:2]
    
    return customer_state != company_state


async def get_company_gst():
    """Get company GST number from settings"""
    settings = await db.company_settings.find_one({}, {"_id": 0})
    if settings and settings.get('gst_number'):
        return settings['gst_number']
    return "29AAAAA1234A1Z5"  # Default


async def generate_quotation_number():
    """Generate quotation number in format QT-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    last_quotation = await db.quotations.find_one(
        {"quotation_number": {"$regex": f"^QT-{date_str}"}},
        sort=[("quotation_number", -1)]
    )
    
    if last_quotation:
        last_num = int(last_quotation["quotation_number"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"QT-{date_str}-{new_num:04d}"


async def generate_sales_order_number():
    """Generate sales order number in format SO-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    last_so = await db.sales_orders.find_one(
        {"so_number": {"$regex": f"^SO-{date_str}"}},
        sort=[("so_number", -1)]
    )
    
    if last_so:
        last_num = int(last_so["so_number"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"SO-{date_str}-{new_num:04d}"


async def generate_buyer_order_number():
    """Generate buyer order number in format BO-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    last_bo = await db.sales_orders.find_one(
        {"buyer_order_no": {"$regex": f"^BO-{date_str}"}},
        sort=[("buyer_order_no", -1)]
    )
    
    if last_bo:
        last_num = int(last_bo["buyer_order_no"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"BO-{date_str}-{new_num:04d}"


async def generate_challan_number():
    """Generate delivery challan number in format DC-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    last_challan = await db.delivery_challans.find_one(
        {"challan_number": {"$regex": f"^DC-{date_str}"}},
        sort=[("challan_number", -1)]
    )
    
    if last_challan:
        last_num = int(last_challan["challan_number"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"DC-{date_str}-{new_num:04d}"


async def generate_credit_note_number():
    """Generate credit note number in format CN-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    last_cn = await db.credit_notes.find_one(
        {"credit_note_number": {"$regex": f"^CN-{date_str}"}},
        sort=[("credit_note_number", -1)]
    )
    
    if last_cn:
        last_num = int(last_cn["credit_note_number"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"CN-{date_str}-{new_num:04d}"


async def generate_payment_number(payment_type: str):
    """Generate payment number based on type"""
    today = datetime.now(timezone.utc)
    year = today.year
    
    if payment_type == "receive":
        prefix = "CUST.IN"
        collection = db.payments
        query = {"payment_type": "receive"}
    else:  # pay
        prefix = "SUPP.OUT"
        collection = db.payments
        query = {"payment_type": "pay"}
    
    last_payment = await collection.find_one(
        query,
        sort=[("payment_number", -1)]
    )
    
    if last_payment:
        try:
            last_num = int(last_payment["payment_number"].split("/")[-1])
            new_num = last_num + 1
        except:
            new_num = 1
    else:
        new_num = 1
    
    return f"{prefix}/{year}/{new_num:05d}"


async def generate_journal_entry_number():
    """Generate journal entry number in format JE-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    last_entry = await db.journal_entries.find_one(
        {"entry_number": {"$regex": f"^JE-{date_str}"}},
        sort=[("entry_number", -1)]
    )
    
    if last_entry:
        last_num = int(last_entry["entry_number"].split("-")[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"JE-{date_str}-{new_num:04d}"


async def update_invoice_payment_status_from_payment(invoice_id: str, invoice_type: str):
    """Update invoice payment status based on total payments received"""
    # Determine collection
    if invoice_type == "sales_invoice":
        collection = db.invoices
    else:  # purchase_invoice
        collection = db.purchase_invoices
    
    # Get invoice
    invoice = await collection.find_one({"id": invoice_id})
    if not invoice:
        return
    
    grand_total = invoice.get('grand_total', 0)
    
    # Get all payments for this invoice
    payments = await db.payments.find({}, {"_id": 0}).to_list(10000)
    
    total_paid = 0
    for payment in payments:
        for allocation in payment.get('allocations', []):
            if allocation['invoice_id'] == invoice_id:
                total_paid += allocation['allocated_amount']
    
    # Determine payment status
    if total_paid == 0:
        payment_status = "unpaid"
    elif total_paid >= grand_total:
        payment_status = "paid"
    else:
        payment_status = "partial"
    
    # Update invoice
    await collection.update_one(
        {"id": invoice_id},
        {"$set": {"payment_status": payment_status}}
    )


async def restore_stock_on_return(items):
    """Restore stock when credit note is created (product return)"""
    for item in items:
        await db.products.update_one(
            {"id": item.product_id},
            {"$inc": {"stock_quantity": item.quantity}}
        )


async def calculate_document_totals(items, customer_gst, overall_discount_type="percentage", overall_discount_value=0.0):
    """Calculate totals for sales documents (Quotation, Sales Order, Invoice)"""
    # Calculate item-level totals
    subtotal = 0
    total_discount = 0
    
    for item in items:
        item_total = item.quantity * item.price
        discount_amount = item_total * (item.discount_percent / 100)
        
        subtotal += item_total
        total_discount += discount_amount
    
    # Calculate after item discounts
    amount_after_item_discount = subtotal - total_discount
    
    # Calculate overall discount
    overall_discount_amount = 0.0
    if overall_discount_value > 0:
        if overall_discount_type == "percentage":
            overall_discount_amount = amount_after_item_discount * (overall_discount_value / 100)
        else:  # amount
            overall_discount_amount = overall_discount_value
    
    # Calculate taxable amount
    taxable_amount = amount_after_item_discount - overall_discount_amount
    
    # Get company GST for interstate detection
    company_gst = await get_company_gst()
    
    # Detect if interstate
    is_interstate = detect_interstate(customer_gst, company_gst)
    
    # Calculate GST
    total_gst = 0
    for item in items:
        item_total = item.quantity * item.price
        item_discount = item_total * (item.discount_percent / 100)
        item_taxable = item_total - item_discount
        
        # Apply overall discount proportionally
        if amount_after_item_discount > 0:
            item_share = item_taxable / amount_after_item_discount
            item_overall_discount = overall_discount_amount * item_share
            item_final_taxable = item_taxable - item_overall_discount
        else:
            item_final_taxable = item_taxable
        
        item_gst = item_final_taxable * (item.gst_rate / 100)
        total_gst += item_gst
    
    # Split GST based on interstate or intrastate
    if is_interstate:
        igst_amount = total_gst
        cgst_amount = 0.0
        sgst_amount = 0.0
    else:
        igst_amount = 0.0
        cgst_amount = total_gst / 2
        sgst_amount = total_gst / 2
    
    grand_total = taxable_amount + total_gst
    
    return {
        "subtotal": subtotal,
        "total_discount": total_discount,
        "overall_discount_amount": overall_discount_amount,
        "taxable_amount": taxable_amount,
        "is_interstate": is_interstate,
        "cgst_amount": cgst_amount,
        "sgst_amount": sgst_amount,
        "igst_amount": igst_amount,
        "total_gst": total_gst,
        "grand_total": grand_total
    }




async def generate_production_order_number():
    """Generate production order number in format PRO-YYYYMMDD-XXXX"""
    today = datetime.now(timezone.utc)
    date_str = today.strftime("%Y%m%d")
    
    # Find the last production order for today
    last_order = await db.production_orders.find_one(
        {"order_number": {"$regex": f"^PRO-{date_str}"}},
        sort=[("order_number", -1)]
    )
    
    if last_order:
        # Extract sequence number and increment
        last_seq = int(last_order["order_number"].split("-")[-1])
        new_seq = last_seq + 1
    else:
        new_seq = 1
    
    return f"PRO-{date_str}-{new_seq:04d}"


# ========== COMPANY SETTINGS ROUTES ==========

@api_router.post("/company-settings", response_model=CompanySettings)
async def create_or_update_company_settings(settings: CompanySettingsCreate):
    """Create or update company settings (singleton)"""
    # Check if settings already exist
    existing = await db.company_settings.find_one({}, {"_id": 0})
    
    if existing:
        # Update existing
        update_data = settings.model_dump()
        update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
        
        await db.company_settings.update_one(
            {"id": existing['id']},
            {"$set": update_data}
        )
        
        updated_settings = await db.company_settings.find_one({"id": existing['id']}, {"_id": 0})
        if isinstance(updated_settings.get('created_at'), str):
            updated_settings['created_at'] = datetime.fromisoformat(updated_settings['created_at'])
        if isinstance(updated_settings.get('updated_at'), str):
            updated_settings['updated_at'] = datetime.fromisoformat(updated_settings['updated_at'])
        return updated_settings
    else:
        # Create new
        settings_obj = CompanySettings(**settings.model_dump())
        doc = settings_obj.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        
        await db.company_settings.insert_one(doc)
        return settings_obj

@api_router.get("/company-settings", response_model=CompanySettings)
async def get_company_settings():
    """Get company settings"""
    settings = await db.company_settings.find_one({}, {"_id": 0})
    
    if not settings:
        # Return default settings if none exist
        return CompanySettings()
    
    if isinstance(settings.get('created_at'), str):
        settings['created_at'] = datetime.fromisoformat(settings['created_at'])
    if isinstance(settings.get('updated_at'), str):
        settings['updated_at'] = datetime.fromisoformat(settings['updated_at'])
    
    return settings


# ========== PRODUCT ROUTES ==========

@api_router.post("/products", response_model=Product)
async def create_product(product: ProductCreate):
    product_obj = Product(**product.model_dump())
    doc = product_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.products.insert_one(doc)
    return product_obj

@api_router.get("/products", response_model=List[Product])
async def get_products():
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    for product in products:
        if isinstance(product.get('created_at'), str):
            product['created_at'] = datetime.fromisoformat(product['created_at'])
    return products

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if isinstance(product.get('created_at'), str):
        product['created_at'] = datetime.fromisoformat(product['created_at'])
    return product

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, product_update: ProductUpdate):
    update_data = {k: v for k, v in product_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.products.update_one(
        {"id": product_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    updated_product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if isinstance(updated_product.get('created_at'), str):
        updated_product['created_at'] = datetime.fromisoformat(updated_product['created_at'])
    return updated_product

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted successfully"}


# ========== CUSTOMER ROUTES ==========

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer: CustomerCreate):
    customer_obj = Customer(**customer.model_dump())
    doc = customer_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.customers.insert_one(doc)
    return customer_obj

@api_router.get("/customers", response_model=List[Customer])
async def get_customers():
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    for customer in customers:
        if isinstance(customer.get('created_at'), str):
            customer['created_at'] = datetime.fromisoformat(customer['created_at'])
    return customers

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    if isinstance(customer.get('created_at'), str):
        customer['created_at'] = datetime.fromisoformat(customer['created_at'])
    return customer

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer_update: CustomerUpdate):
    update_data = {k: v for k, v in customer_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    updated_customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if isinstance(updated_customer.get('created_at'), str):
        updated_customer['created_at'] = datetime.fromisoformat(updated_customer['created_at'])
    return updated_customer

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str):
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted successfully"}


# ========== SUPPLIER ROUTES ==========

@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(supplier: SupplierCreate):
    supplier_obj = Supplier(**supplier.model_dump())
    doc = supplier_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.suppliers.insert_one(doc)
    return supplier_obj

@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers():
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    for supplier in suppliers:
        if isinstance(supplier.get('created_at'), str):
            supplier['created_at'] = datetime.fromisoformat(supplier['created_at'])
    return suppliers

@api_router.get("/suppliers/{supplier_id}", response_model=Supplier)
async def get_supplier(supplier_id: str):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    if isinstance(supplier.get('created_at'), str):
        supplier['created_at'] = datetime.fromisoformat(supplier['created_at'])
    return supplier

@api_router.put("/suppliers/{supplier_id}", response_model=Supplier)
async def update_supplier(supplier_id: str, supplier_update: SupplierUpdate):
    update_data = {k: v for k, v in supplier_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    updated_supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if isinstance(updated_supplier.get('created_at'), str):
        updated_supplier['created_at'] = datetime.fromisoformat(updated_supplier['created_at'])
    return updated_supplier

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Supplier deleted successfully"}


# ========== RAW MATERIAL ROUTES ==========

@api_router.post("/raw-materials", response_model=RawMaterial)
async def create_raw_material(raw_material: RawMaterialCreate):
    raw_material_obj = RawMaterial(**raw_material.model_dump())
    doc = raw_material_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.raw_materials.insert_one(doc)
    return raw_material_obj

@api_router.get("/raw-materials", response_model=List[RawMaterial])
async def get_raw_materials():
    raw_materials = await db.raw_materials.find({}, {"_id": 0}).to_list(1000)
    for raw_material in raw_materials:
        if isinstance(raw_material.get('created_at'), str):
            raw_material['created_at'] = datetime.fromisoformat(raw_material['created_at'])
    return raw_materials

@api_router.get("/raw-materials/{raw_material_id}", response_model=RawMaterial)
async def get_raw_material(raw_material_id: str):
    raw_material = await db.raw_materials.find_one({"id": raw_material_id}, {"_id": 0})
    if not raw_material:
        raise HTTPException(status_code=404, detail="Raw material not found")
    if isinstance(raw_material.get('created_at'), str):
        raw_material['created_at'] = datetime.fromisoformat(raw_material['created_at'])
    return raw_material

@api_router.put("/raw-materials/{raw_material_id}", response_model=RawMaterial)
async def update_raw_material(raw_material_id: str, raw_material_update: RawMaterialUpdate):
    update_data = {k: v for k, v in raw_material_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.raw_materials.update_one(
        {"id": raw_material_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    
    updated_raw_material = await db.raw_materials.find_one({"id": raw_material_id}, {"_id": 0})
    if isinstance(updated_raw_material.get('created_at'), str):
        updated_raw_material['created_at'] = datetime.fromisoformat(updated_raw_material['created_at'])
    return updated_raw_material

@api_router.delete("/raw-materials/{raw_material_id}")
async def delete_raw_material(raw_material_id: str):
    result = await db.raw_materials.delete_one({"id": raw_material_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    return {"message": "Raw material deleted successfully"}


# ========== PACKING MATERIAL ROUTES ==========

@api_router.post("/packing-materials", response_model=PackingMaterial)
async def create_packing_material(packing_material: PackingMaterialCreate):
    packing_material_obj = PackingMaterial(**packing_material.model_dump())
    doc = packing_material_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.packing_materials.insert_one(doc)
    return packing_material_obj

@api_router.get("/packing-materials", response_model=List[PackingMaterial])
async def get_packing_materials():
    packing_materials = await db.packing_materials.find({}, {"_id": 0}).to_list(1000)
    for packing_material in packing_materials:
        if isinstance(packing_material.get('created_at'), str):
            packing_material['created_at'] = datetime.fromisoformat(packing_material['created_at'])
    return packing_materials

@api_router.get("/packing-materials/{packing_material_id}", response_model=PackingMaterial)
async def get_packing_material(packing_material_id: str):
    packing_material = await db.packing_materials.find_one({"id": packing_material_id}, {"_id": 0})
    if not packing_material:
        raise HTTPException(status_code=404, detail="Packing material not found")
    if isinstance(packing_material.get('created_at'), str):
        packing_material['created_at'] = datetime.fromisoformat(packing_material['created_at'])
    return packing_material

@api_router.put("/packing-materials/{packing_material_id}", response_model=PackingMaterial)
async def update_packing_material(packing_material_id: str, packing_material_update: PackingMaterialUpdate):
    update_data = {k: v for k, v in packing_material_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.packing_materials.update_one(
        {"id": packing_material_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Packing material not found")
    
    updated_packing_material = await db.packing_materials.find_one({"id": packing_material_id}, {"_id": 0})
    if isinstance(updated_packing_material.get('created_at'), str):
        updated_packing_material['created_at'] = datetime.fromisoformat(updated_packing_material['created_at'])
    return updated_packing_material

@api_router.delete("/packing-materials/{packing_material_id}")
async def delete_packing_material(packing_material_id: str):
    result = await db.packing_materials.delete_one({"id": packing_material_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Packing material not found")
    return {"message": "Packing material deleted successfully"}




# ========== PURCHASE REQUEST ROUTES ==========

@api_router.post("/purchase-requests", response_model=PurchaseRequest)
async def create_purchase_request(request_data: PurchaseRequestCreate):
    """Inventory manager creates purchase request"""
    # Generate request number
    count = await db.purchase_requests.count_documents({})
    request_number = f"PR-{datetime.now().strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"
    
    # Create purchase request
    request_obj = PurchaseRequest(
        request_number=request_number,
        **request_data.model_dump()
    )
    
    doc = request_obj.model_dump()
    doc['request_date'] = doc['request_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.purchase_requests.insert_one(doc)
    
    return request_obj


@api_router.get("/purchase-requests/pending")
async def get_pending_purchase_requests():
    """Get all pending purchase requests for purchase manager approval"""
    requests = await db.purchase_requests.find(
        {"status": "pending"},
        {"_id": 0}
    ).to_list(length=None)
    return requests


@api_router.get("/purchase-requests")
async def get_all_purchase_requests():
    """Get all purchase requests"""
    requests = await db.purchase_requests.find(
        {},
        {"_id": 0}
    ).sort("created_at", -1).to_list(length=None)
    return requests


@api_router.post("/purchase-requests/{request_id}/approve")
async def approve_purchase_request(request_id: str, approval_data: dict):
    """Purchase manager approves purchase request - initial approval"""
    request = await db.purchase_requests.find_one({"id": request_id}, {"_id": 0})
    
    if not request:
        raise HTTPException(status_code=404, detail="Purchase request not found")
    
    if request.get('status') != 'pending':
        raise HTTPException(status_code=400, detail="Request is not pending")
    
    approved_by = approval_data.get("approved_by", "Purchase Manager")
    
    # Update purchase request to approved
    await db.purchase_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": approved_by,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Purchase request approved by {approved_by}"}


@api_router.post("/purchase-requests/{request_id}/quote")
async def quote_purchase_request(request_id: str, quote_data: dict):
    """Purchase manager adds costs, supplier and sends to finance for approval"""
    request = await db.purchase_requests.find_one({"id": request_id}, {"_id": 0})
    
    if not request:
        raise HTTPException(status_code=404, detail="Purchase request not found")
    
    if request.get('status') != 'approved':
        raise HTTPException(status_code=400, detail="Request must be approved first")
    
    # Update with pricing and supplier information
    items_with_costs = quote_data.get("items", [])
    supplier_id = quote_data.get("supplier_id")
    supplier_name = quote_data.get("supplier_name")
    total_cost = sum(item.get("quantity", 0) * item.get("unit_cost", 0) for item in items_with_costs)
    
    await db.purchase_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "quoted",
            "items": items_with_costs,
            "supplier_id": supplier_id,
            "supplier_name": supplier_name,
            "total_cost": total_cost,
            "quoted_by": quote_data.get("quoted_by", "Purchase Manager"),
            "quoted_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Quote submitted for finance approval", "total_cost": total_cost}


@api_router.get("/purchase-requests/finance-pending/list")
async def get_finance_pending_requests():
    """Get purchase requests pending finance approval (quoted status)"""
    requests = await db.purchase_requests.find(
        {"status": "quoted"},
        {"_id": 0}
    ).sort("quoted_at", -1).to_list(1000)
    return requests


@api_router.post("/purchase-requests/{request_id}/finance-approve")
async def finance_approve_purchase_request(request_id: str, approval_data: dict):
    """Finance manager approves purchase request and creates purchase order"""
    request = await db.purchase_requests.find_one({"id": request_id}, {"_id": 0})
    
    if not request:
        raise HTTPException(status_code=404, detail="Purchase request not found")
    
    if request.get('status') != 'quoted':
        raise HTTPException(status_code=400, detail="Request must be quoted first")
    
    approved_by = approval_data.get("approved_by", "Finance Manager")
    
    # Update purchase request to finance_approved
    await db.purchase_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "finance_approved",
            "finance_approved_by": approved_by,
            "finance_approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Auto-create Purchase Order
    po_number = await generate_po_number()
    
    po_data = {
        "id": str(uuid4()),
        "po_number": po_number,
        "supplier_id": request.get("supplier_id"),
        "supplier_name": request.get("supplier_name"),
        "po_date": datetime.now(timezone.utc).isoformat(),
        "items": request.get("items", []),
        "subtotal": request.get("total_cost", 0),
        "tax_amount": 0,
        "total_amount": request.get("total_cost", 0),
        "status": "pending",
        "payment_terms": "As per agreement",
        "delivery_date": None,
        "notes": f"Auto-generated from Purchase Request: {request.get('request_number')}",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": approved_by,
        "purchase_request_id": request_id
    }
    
    await db.purchase_orders.insert_one(po_data)
    
    # Update purchase request with PO reference
    await db.purchase_requests.update_one(
        {"id": request_id},
        {"$set": {"purchase_order_id": po_data["id"], "purchase_order_number": po_number}}
    )
    
    return {
        "message": f"Purchase request approved by {approved_by} and Purchase Order {po_number} created",
        "po_number": po_number,
        "po_id": po_data["id"]
    }


@api_router.post("/purchase-requests/{request_id}/finance-reject")
async def finance_reject_purchase_request(request_id: str, rejection_data: dict):
    """Finance manager rejects purchase request"""
    request = await db.purchase_requests.find_one({"id": request_id}, {"_id": 0})
    
    if not request:
        raise HTTPException(status_code=404, detail="Purchase request not found")
    
    if request.get('status') != 'quoted':
        raise HTTPException(status_code=400, detail="Request must be quoted first")
    
    rejection_reason = rejection_data.get("reason", "Budget constraints")
    rejected_by = rejection_data.get("rejected_by", "Finance Manager")
    
    await db.purchase_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "finance_rejected",
            "finance_rejection_reason": rejection_reason,
            "finance_rejected_by": rejected_by,
            "finance_rejected_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Purchase request rejected by finance: {rejection_reason}"}


@api_router.post("/purchase-requests/{request_id}/reject")
async def reject_purchase_request(request_id: str, rejection_data: dict):
    """Purchase manager rejects purchase request"""
    request = await db.purchase_requests.find_one({"id": request_id}, {"_id": 0})
    
    if not request:
        raise HTTPException(status_code=404, detail="Purchase request not found")
    
    if request.get('status') != 'pending':
        raise HTTPException(status_code=400, detail="Request is not pending")
    
    rejection_reason = rejection_data.get("reason", "No reason provided")
    rejected_by = rejection_data.get("rejected_by", "Purchase Manager")
    
    await db.purchase_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "rejection_reason": rejection_reason,
            "approved_by": rejected_by,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Purchase request rejected: {rejection_reason}"}


# ========== STOCK INWARD ROUTES ==========

@api_router.post("/stock-inward")
async def create_stock_inward(stock_data: StockInwardCreate):
    """Add stock to raw material or packing material"""
    # Determine collection
    collection_name = "raw_materials" if stock_data.material_type == "raw_material" else "packing_materials"
    collection = db[collection_name]
    
    # Get material
    material = await collection.find_one({"id": stock_data.material_id}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    
    # Update stock quantity
    new_quantity = material.get("stock_quantity", 0) + stock_data.quantity_added
    await collection.update_one(
        {"id": stock_data.material_id},
        {"$set": {"stock_quantity": new_quantity}}
    )
    
    # Save stock inward record
    stock_inward = StockInward(
        material_type=stock_data.material_type,
        material_id=stock_data.material_id,
        material_name=material.get("name", ""),
        quantity_added=stock_data.quantity_added,
        unit=material.get("unit", "")
    )
    
    doc = stock_inward.model_dump()
    doc['added_date'] = doc['added_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.stock_inward.insert_one(doc)
    
    return {
        "message": "Stock added successfully",
        "material_name": material.get("name"),
        "quantity_added": stock_data.quantity_added,
        "new_stock_quantity": new_quantity,
        "unit": material.get("unit")
    }


@api_router.get("/stock-inward")
async def get_stock_inward():
    """Get all stock inward entries"""
    entries = await db.stock_inward.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return entries



# ========== SUPPLIER PRICE ROUTES ==========

@api_router.post("/supplier-prices")
async def create_supplier_price(price_data: SupplierPriceCreate):
    """Add or update supplier price for a material"""
    # Get supplier details
    supplier = await db.suppliers.find_one({"id": price_data.supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Get material details
    collection_name = "raw_materials" if price_data.material_type == "raw_material" else "packing_materials"
    collection = db[collection_name]
    material = await collection.find_one({"id": price_data.material_id}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Material not found")
    
    # Check if price already exists for this supplier-material combination
    existing = await db.supplier_prices.find_one({
        "supplier_id": price_data.supplier_id,
        "material_id": price_data.material_id
    }, {"_id": 0})
    
    if existing:
        # Update existing price
        update_data = {
            "price": price_data.price,
            "lead_time_days": price_data.lead_time_days,
            "minimum_order_qty": price_data.minimum_order_qty,
            "notes": price_data.notes,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.supplier_prices.update_one(
            {"id": existing['id']},
            {"$set": update_data}
        )
        
        return {"message": "Supplier price updated successfully", "id": existing['id']}
    else:
        # Create new price entry
        supplier_price = SupplierPrice(
            supplier_id=price_data.supplier_id,
            supplier_name=supplier.get("name", ""),
            material_type=price_data.material_type,
            material_id=price_data.material_id,
            material_name=material.get("name", ""),
            unit=material.get("unit", ""),
            price=price_data.price,
            lead_time_days=price_data.lead_time_days,
            minimum_order_qty=price_data.minimum_order_qty,
            notes=price_data.notes
        )
        
        doc = supplier_price.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        
        await db.supplier_prices.insert_one(doc)
        
        return {"message": "Supplier price added successfully", "id": supplier_price.id}


@api_router.get("/supplier-prices")
async def get_supplier_prices(
    supplier_id: Optional[str] = None,
    material_id: Optional[str] = None,
    material_type: Optional[str] = None
):
    """Get all supplier prices with optional filters"""
    query = {}
    
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    if material_id:
        query["material_id"] = material_id
    
    if material_type:
        query["material_type"] = material_type
    
    prices = await db.supplier_prices.find(query, {"_id": 0}).to_list(1000)
    return prices


@api_router.get("/supplier-prices/comparison")
async def get_price_comparison(material_type: Optional[str] = None):
    """Get price comparison across all suppliers and materials"""
    query = {}
    if material_type:
        query["material_type"] = material_type
    
    # Get all supplier prices
    all_prices = await db.supplier_prices.find(query, {"_id": 0}).to_list(10000)
    
    # Group by material
    materials_map = {}
    
    for price in all_prices:
        material_id = price['material_id']
        
        if material_id not in materials_map:
            materials_map[material_id] = {
                "material_id": material_id,
                "material_name": price['material_name'],
                "material_type": price['material_type'],
                "unit": price['unit'],
                "suppliers": []
            }
        
        materials_map[material_id]['suppliers'].append({
            "supplier_id": price['supplier_id'],
            "supplier_name": price['supplier_name'],
            "price": price['price'],
            "lead_time_days": price.get('lead_time_days'),
            "minimum_order_qty": price.get('minimum_order_qty'),
            "notes": price.get('notes', ''),
            "updated_at": price['updated_at']
        })
    
    # Convert to list and find cheapest for each material
    comparison = []
    for material_id, data in materials_map.items():
        # Sort suppliers by price
        data['suppliers'].sort(key=lambda x: x['price'])
        
        # Mark the cheapest
        if data['suppliers']:
            data['cheapest_supplier'] = data['suppliers'][0]['supplier_name']
            data['cheapest_price'] = data['suppliers'][0]['price']
        
        comparison.append(data)
    
    # Sort by material name
    comparison.sort(key=lambda x: x['material_name'])
    
    return comparison


@api_router.get("/supplier-prices/{price_id}")
async def get_supplier_price(price_id: str):
    """Get single supplier price"""
    price = await db.supplier_prices.find_one({"id": price_id}, {"_id": 0})
    if not price:
        raise HTTPException(status_code=404, detail="Supplier price not found")
    return price


@api_router.put("/supplier-prices/{price_id}")
async def update_supplier_price(price_id: str, update_data: SupplierPriceUpdate):
    """Update supplier price"""
    existing = await db.supplier_prices.find_one({"id": price_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Supplier price not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.supplier_prices.update_one(
        {"id": price_id},
        {"$set": update_dict}
    )
    
    updated_price = await db.supplier_prices.find_one({"id": price_id}, {"_id": 0})
    return {"message": "Supplier price updated successfully", "price": updated_price}


@api_router.delete("/supplier-prices/{price_id}")
async def delete_supplier_price(price_id: str):
    """Delete supplier price"""
    result = await db.supplier_prices.delete_one({"id": price_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier price not found")
    
    return {"message": "Supplier price deleted successfully"}


# ========== PURCHASE ORDER ROUTES ==========

@api_router.post("/purchase-orders", response_model=PurchaseOrder)
async def create_purchase_order(po_data: PurchaseOrderCreate):
    # Generate PO number
    po_number = await generate_po_number()
    
    # Calculate totals
    subtotal = 0
    total_gst = 0
    
    for item in po_data.items:
        item_total = item.quantity * item.price
        item_gst = item_total * (item.gst_rate / 100)
        
        subtotal += item_total
        total_gst += item_gst
    
    grand_total = subtotal + total_gst
    
    po_obj = PurchaseOrder(
        po_number=po_number,
        **po_data.model_dump(),
        subtotal=subtotal,
        total_gst=total_gst,
        grand_total=grand_total
    )
    
    doc = po_obj.model_dump()
    doc['po_date'] = doc['po_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.purchase_orders.insert_one(doc)
    return po_obj

@api_router.get("/purchase-orders", response_model=List[PurchaseOrder])
async def get_purchase_orders():
    pos = await db.purchase_orders.find({}, {"_id": 0}).sort("po_date", -1).to_list(1000)
    for po in pos:
        if isinstance(po.get('po_date'), str):
            po['po_date'] = datetime.fromisoformat(po['po_date'])
        if isinstance(po.get('created_at'), str):
            po['created_at'] = datetime.fromisoformat(po['created_at'])
    return pos

@api_router.get("/purchase-orders/{po_id}", response_model=PurchaseOrder)
async def get_purchase_order(po_id: str):
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    if isinstance(po.get('po_date'), str):
        po['po_date'] = datetime.fromisoformat(po['po_date'])
    if isinstance(po.get('created_at'), str):
        po['created_at'] = datetime.fromisoformat(po['created_at'])
    return po

@api_router.put("/purchase-orders/{po_id}/status")
async def update_po_status(po_id: str, status_data: dict):
    status = status_data.get("status")
    if status not in ["draft", "confirmed", "received", "cancelled"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"status": status}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    return {"message": "Status updated successfully"}

@api_router.delete("/purchase-orders/{po_id}")
async def delete_purchase_order(po_id: str):
    """Delete a purchase order"""
    result = await db.purchase_orders.delete_one({"id": po_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    return {"message": "Purchase order deleted successfully"}


# ========== PURCHASE INVOICE ROUTES ==========

@api_router.post("/purchase-invoices", response_model=PurchaseInvoice)
async def create_purchase_invoice(invoice_data: PurchaseInvoiceCreate):
    # Generate invoice number
    invoice_number = await generate_purchase_invoice_number()
    
    # Calculate totals
    subtotal = 0
    total_discount = 0
    taxable_amount = 0
    total_gst = 0
    
    for item in invoice_data.items:
        item_total = item.quantity * item.price
        discount_amount = item_total * (item.discount_percent / 100)
        item_taxable = item_total - discount_amount
        item_gst = item_taxable * (item.gst_rate / 100)
        
        subtotal += item_total
        total_discount += discount_amount
        taxable_amount += item_taxable
        total_gst += item_gst
    
    cgst_amount = total_gst / 2
    sgst_amount = total_gst / 2
    grand_total = taxable_amount + total_gst
    
    invoice_obj = PurchaseInvoice(
        invoice_number=invoice_number,
        **invoice_data.model_dump(),
        subtotal=subtotal,
        total_discount=total_discount,
        taxable_amount=taxable_amount,
        cgst_amount=cgst_amount,
        sgst_amount=sgst_amount,
        total_gst=total_gst,
        grand_total=grand_total,
        stock_updated=True
    )
    
    # Update stock for each item
    await update_stock_on_purchase(invoice_data.items)
    
    doc = invoice_obj.model_dump()
    doc['invoice_date'] = doc['invoice_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.purchase_invoices.insert_one(doc)
    return invoice_obj

@api_router.get("/purchase-invoices", response_model=List[PurchaseInvoice])
async def get_purchase_invoices():
    invoices = await db.purchase_invoices.find({}, {"_id": 0}).sort("invoice_date", -1).to_list(1000)
    for invoice in invoices:
        if isinstance(invoice.get('invoice_date'), str):
            invoice['invoice_date'] = datetime.fromisoformat(invoice['invoice_date'])
        if isinstance(invoice.get('created_at'), str):
            invoice['created_at'] = datetime.fromisoformat(invoice['created_at'])
    return invoices

@api_router.get("/purchase-invoices/{invoice_id}", response_model=PurchaseInvoice)
async def get_purchase_invoice(invoice_id: str):
    invoice = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Purchase invoice not found")
    if isinstance(invoice.get('invoice_date'), str):
        invoice['invoice_date'] = datetime.fromisoformat(invoice['invoice_date'])
    if isinstance(invoice.get('created_at'), str):
        invoice['created_at'] = datetime.fromisoformat(invoice['created_at'])
    return invoice

@api_router.put("/purchase-invoices/{invoice_id}/payment-status")
async def update_purchase_payment_status(invoice_id: str, status: dict):
    payment_status = status.get("payment_status")
    if payment_status not in ["unpaid", "partial", "paid"]:
        raise HTTPException(status_code=400, detail="Invalid payment status")
    
    result = await db.purchase_invoices.update_one(
        {"id": invoice_id},
        {"$set": {"payment_status": payment_status}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Purchase invoice not found")
    
    return {"message": "Payment status updated successfully"}




# ========== FINANCIAL TRANSACTIONS ROUTES ==========

@api_router.post("/financial-transactions")
async def create_financial_transaction(transaction_data: FinancialTransactionCreate):
    """Create a financial transaction"""
    # Prepare data
    data_dict = transaction_data.model_dump()
    if data_dict.get('transaction_date') is None:
        data_dict['transaction_date'] = datetime.now(timezone.utc)
    
    transaction = FinancialTransaction(**data_dict)
    
    doc = transaction.model_dump()
    doc['transaction_date'] = doc['transaction_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.financial_transactions.insert_one(doc)
    
    return {"message": "Financial transaction created successfully", "transaction": transaction.model_dump()}


@api_router.get("/financial-transactions")
async def get_financial_transactions(
    transaction_type: Optional[str] = None,
    category: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Get all financial transactions with filters"""
    query = {}
    
    if transaction_type:
        query["transaction_type"] = transaction_type
    
    if category:
        query["category"] = category
    
    if start_date:
        query["transaction_date"] = {"$gte": start_date}
    
    if end_date:
        if "transaction_date" in query:
            query["transaction_date"]["$lte"] = end_date
        else:
            query["transaction_date"] = {"$lte": end_date}
    
    transactions = await db.financial_transactions.find(query, {"_id": 0}).sort("transaction_date", -1).to_list(1000)
    return transactions


@api_router.get("/balance-sheet")
async def get_balance_sheet():
    """Get comprehensive balance sheet with all financial data"""
    
    # Get all financial transactions
    all_transactions = await db.financial_transactions.find({}, {"_id": 0}).to_list(10000)
    
    # Calculate payments
    payments_received = sum(t.get('amount', 0) for t in all_transactions if t.get('transaction_type') == 'payment_received')
    payments_made = sum(t.get('amount', 0) for t in all_transactions if t.get('transaction_type') == 'payment_made')
    
    # Get purchase invoices
    purchase_invoices = await db.purchase_invoices.find({}, {"_id": 0}).to_list(10000)
    total_purchases = sum(pi.get('total_amount', 0) for pi in purchase_invoices)
    paid_purchases = sum(pi.get('paid_amount', 0) for pi in purchase_invoices)
    outstanding_payables = total_purchases - paid_purchases
    
    # Get sales invoices
    sales_invoices = await db.invoices.find({}, {"_id": 0}).to_list(10000)
    total_sales = sum(inv.get('total_amount', 0) for inv in sales_invoices)
    paid_sales = sum(inv.get('paid_amount', 0) for inv in sales_invoices)
    outstanding_receivables = total_sales - paid_sales
    
    # Calculate inventory value
    raw_materials = await db.raw_materials.find({}, {"_id": 0}).to_list(1000)
    raw_materials_value = sum(
        (rm.get('stock_quantity', 0) * rm.get('purchase_price', 0)) 
        for rm in raw_materials
    )
    
    packing_materials = await db.packing_materials.find({}, {"_id": 0}).to_list(1000)
    packing_materials_value = sum(
        (pm.get('stock_quantity', 0) * pm.get('purchase_price', 0)) 
        for pm in packing_materials
    )
    
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    finished_goods_value = sum(
        (p.get('stock_quantity', 0) * p.get('selling_price', 0)) 
        for p in products
    )
    
    total_inventory_value = raw_materials_value + packing_materials_value + finished_goods_value
    
    # Calculate expenses (from financial transactions)
    expenses = sum(t.get('amount', 0) for t in all_transactions if t.get('category') == 'expense')
    
    # Calculate current cash balance
    cash_balance = payments_received - payments_made
    
    # Calculate net worth
    total_assets = cash_balance + total_inventory_value + outstanding_receivables
    total_liabilities = outstanding_payables


# ========== AUTHENTICATION ROUTES ==========

@api_router.post("/auth/register", response_model=dict)
async def register_user(user_data: UserCreate):
    """Register a new user (admin only in production)"""
    # Check if username already exists
    existing_user = await db.users.find_one({"username": user_data.username}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Create user
    hashed_password = get_password_hash(user_data.password)
    user = User(
        username=user_data.username,
        email=user_data.email,
        hashed_password=hashed_password,
        full_name=user_data.full_name,
        role=user_data.role
    )
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    return {"message": "User created successfully", "username": user.username, "role": user.role}


@api_router.post("/auth/login", response_model=Token)
async def login(user_credentials: UserLogin):
    """Login and get access token"""
    # Find user
    user = await db.users.find_one({"username": user_credentials.username}, {"_id": 0})
    
    if not user:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    # Verify password
    if not verify_password(user_credentials.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="User account is disabled")
    
    # Create access token
    access_token = create_access_token(
        data={"sub": user["username"], "role": user["role"]}
    )
    
    # Return token and user info (without password)
    user_info = {
        "id": user["id"],
        "username": user["username"],
        "email": user.get("email", ""),
        "full_name": user["full_name"],
        "role": user["role"]
    }
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user_info
    }


@api_router.get("/auth/me")
async def get_current_user(token: str):
    """Get current user from token"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = await db.users.find_one({"username": username}, {"_id": 0, "hashed_password": 0})
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    
    return user


@api_router.get("/users")
async def get_all_users():
    """Get all users (admin only)"""
    users = await db.users.find({}, {"_id": 0, "hashed_password": 0}).to_list(100)
    return users


@api_router.post("/auth/init-admin")
async def initialize_admin():
    """Initialize default admin user - run this once"""
    # Check if any admin exists
    existing_admin = await db.users.find_one({"role": "admin"}, {"_id": 0})
    if existing_admin:
        return {"message": "Admin already exists", "username": existing_admin["username"]}
    
    # Create default admin
    admin_password = "admin123"  # Change this immediately after first login!
    hashed_password = get_password_hash(admin_password)
    
    admin = User(
        username="admin",
        email="admin@nectarasia.com",
        hashed_password=hashed_password,
        full_name="System Administrator",
        role="admin"
    )
    
    doc = admin.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    return {
        "message": "Admin user created successfully",
        "username": "admin",
        "password": admin_password,
        "warning": "⚠️ CHANGE THIS PASSWORD IMMEDIATELY!"
    }

    net_worth = total_assets - total_liabilities
    
    # Calculate profit/loss
    gross_revenue = total_sales
    total_expenses = total_purchases + expenses
    net_profit = gross_revenue - total_expenses
    
    return {
        "summary": {
            "cash_balance": round(cash_balance, 2),
            "total_assets": round(total_assets, 2),
            "total_liabilities": round(total_liabilities, 2),
            "net_worth": round(net_worth, 2),
            "net_profit": round(net_profit, 2)
        },
        "assets": {
            "cash": round(cash_balance, 2),
            "inventory": {
                "raw_materials": round(raw_materials_value, 2),
                "packing_materials": round(packing_materials_value, 2),
                "finished_goods": round(finished_goods_value, 2),
                "total": round(total_inventory_value, 2)
            },
            "accounts_receivable": round(outstanding_receivables, 2),
            "total_assets": round(total_assets, 2)
        },
        "liabilities": {
            "accounts_payable": round(outstanding_payables, 2),
            "total_liabilities": round(total_liabilities, 2)
        },
        "income_statement": {
            "revenue": {
                "sales": round(total_sales, 2),
                "payments_received": round(payments_received, 2)
            },
            "expenses": {
                "purchases": round(total_purchases, 2),
                "payments_made": round(payments_made, 2),
                "other_expenses": round(expenses, 2),
                "total_expenses": round(total_expenses, 2)
            },
            "net_profit": round(net_profit, 2)
        },
        "transactions_summary": {
            "total_transactions": len(all_transactions),
            "payments_received_count": sum(1 for t in all_transactions if t.get('transaction_type') == 'payment_received'),
            "payments_made_count": sum(1 for t in all_transactions if t.get('transaction_type') == 'payment_made')
        }
    }


@api_router.delete("/financial-transactions/{transaction_id}")
async def delete_financial_transaction(transaction_id: str):
    """Delete a financial transaction"""
    result = await db.financial_transactions.delete_one({"id": transaction_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    return {"message": "Transaction deleted successfully"}


# ========== INVOICE ROUTES ==========

@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(invoice_data: InvoiceCreate):
    # Generate invoice number
    invoice_number = await generate_invoice_number()
    
    # Calculate totals using helper function
    totals = await calculate_document_totals(
        invoice_data.items,
        invoice_data.customer_gst,
        invoice_data.overall_discount_type,
        invoice_data.overall_discount_value
    )
    
    invoice_obj = Invoice(
        invoice_number=invoice_number,
        **invoice_data.model_dump(),
        **totals,
        stock_updated=True
    )
    
    # Deduct stock for each item
    await update_stock_on_sale(invoice_data.items)
    
    doc = invoice_obj.model_dump()
    doc['invoice_date'] = doc['invoice_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.invoices.insert_one(doc)
    return invoice_obj

@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices():
    invoices = await db.invoices.find({}, {"_id": 0}).sort("invoice_date", -1).to_list(1000)
    for invoice in invoices:
        if isinstance(invoice.get('invoice_date'), str):
            invoice['invoice_date'] = datetime.fromisoformat(invoice['invoice_date'])
        if isinstance(invoice.get('created_at'), str):
            invoice['created_at'] = datetime.fromisoformat(invoice['created_at'])
    return invoices

@api_router.get("/invoices/{invoice_id}", response_model=Invoice)
async def get_invoice(invoice_id: str):
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if isinstance(invoice.get('invoice_date'), str):
        invoice['invoice_date'] = datetime.fromisoformat(invoice['invoice_date'])
    if isinstance(invoice.get('created_at'), str):
        invoice['created_at'] = datetime.fromisoformat(invoice['created_at'])
    return invoice

@api_router.put("/invoices/{invoice_id}/payment-status")
async def update_payment_status(invoice_id: str, status: dict):
    payment_status = status.get("payment_status")
    if payment_status not in ["unpaid", "partial", "paid"]:
        raise HTTPException(status_code=400, detail="Invalid payment status")
    
    result = await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"payment_status": payment_status}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"message": "Payment status updated successfully"}


@api_router.put("/invoices/{invoice_id}/status")
async def update_invoice_status(invoice_id: str, status_data: dict):
    invoice_status = status_data.get("invoice_status")
    if invoice_status not in ["draft", "confirmed", "dispatched", "delivered"]:
        raise HTTPException(status_code=400, detail="Invalid invoice status")
    
    result = await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"invoice_status": invoice_status}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"message": "Invoice status updated successfully"}


@api_router.delete("/invoices/{invoice_id}/cancel")
async def cancel_invoice(invoice_id: str):
    """Delete invoice and restore stock"""
    # Get invoice
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Restore stock for each item (only if stock was updated)
    if invoice.get('stock_updated', True):  # Default to True for old invoices
        for item in invoice['items']:
            await db.products.update_one(
                {"id": item['product_id']},
                {"$inc": {"stock_quantity": item['quantity']}}
            )
    
    # Delete the invoice completely
    result = await db.invoices.delete_one({"id": invoice_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"message": "Invoice deleted and stock restored successfully"}


# ========== QUOTATION ROUTES ==========

@api_router.post("/quotations", response_model=Quotation)
async def create_quotation(quotation_data: QuotationCreate):
    quotation_number = await generate_quotation_number()
    totals = await calculate_document_totals(
        quotation_data.items, quotation_data.customer_gst,
        quotation_data.overall_discount_type, quotation_data.overall_discount_value
    )
    quotation_obj = Quotation(quotation_number=quotation_number, **quotation_data.model_dump(), **totals)
    doc = quotation_obj.model_dump()
    doc['quotation_date'] = doc['quotation_date'].isoformat()
    if doc.get('valid_until'):
        doc['valid_until'] = doc['valid_until'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.quotations.insert_one(doc)
    return quotation_obj

@api_router.get("/quotations", response_model=List[Quotation])
async def get_quotations():
    quots = await db.quotations.find({}, {"_id": 0}).sort("quotation_date", -1).to_list(1000)
    for q in quots:
        if isinstance(q.get('quotation_date'), str):
            q['quotation_date'] = datetime.fromisoformat(q['quotation_date'])
        if q.get('valid_until') and isinstance(q['valid_until'], str):
            q['valid_until'] = datetime.fromisoformat(q['valid_until'])
        if isinstance(q.get('created_at'), str):
            q['created_at'] = datetime.fromisoformat(q['created_at'])
    return quots

@api_router.get("/quotations/{quotation_id}", response_model=Quotation)
async def get_quotation(quotation_id: str):
    q = await db.quotations.find_one({"id": quotation_id}, {"_id": 0})
    if not q:
        raise HTTPException(status_code=404, detail="Quotation not found")
    if isinstance(q.get('quotation_date'), str):
        q['quotation_date'] = datetime.fromisoformat(q['quotation_date'])
    if q.get('valid_until') and isinstance(q['valid_until'], str):
        q['valid_until'] = datetime.fromisoformat(q['valid_until'])
    if isinstance(q.get('created_at'), str):
        q['created_at'] = datetime.fromisoformat(q['created_at'])
    return q

@api_router.put("/quotations/{quotation_id}/status")
async def update_quotation_status(quotation_id: str, status_data: dict):
    status = status_data.get("quotation_status")
    if status not in ["draft", "sent", "accepted", "rejected", "converted"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    await db.quotations.update_one({"id": quotation_id}, {"$set": {"quotation_status": status}})
    return {"message": "Status updated"}

@api_router.delete("/quotations/{quotation_id}")
async def delete_quotation(quotation_id: str):
    """Delete a quotation"""
    result = await db.quotations.delete_one({"id": quotation_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Quotation not found")
    return {"message": "Quotation deleted successfully"}


# ========== SALES ORDER ROUTES ==========

@api_router.post("/sales-orders", response_model=SalesOrder)
async def create_sales_order(so_data: SalesOrderCreate):
    so_number = await generate_sales_order_number()
    
    # Auto-generate buyer order number if not provided
    buyer_order_no = so_data.buyer_order_no
    if not buyer_order_no or buyer_order_no.strip() == "":
        buyer_order_no = await generate_buyer_order_number()
    
    totals = await calculate_document_totals(
        so_data.items, so_data.customer_gst,
        so_data.overall_discount_type, so_data.overall_discount_value
    )
    
    # Create sales order with auto-generated buyer order number
    so_data_dict = so_data.model_dump()
    so_data_dict['buyer_order_no'] = buyer_order_no
    
    # Set default approval workflow fields
    so_data_dict['approval_status'] = 'pending'
    so_data_dict['approved_by'] = None
    so_data_dict['approved_at'] = None
    so_data_dict['rejection_reason'] = None
    
    so_obj = SalesOrder(so_number=so_number, **so_data_dict, **totals)
    doc = so_obj.model_dump()
    doc['so_date'] = doc['so_date'].isoformat()
    if doc.get('expected_delivery_date'):
        doc['expected_delivery_date'] = doc['expected_delivery_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.sales_orders.insert_one(doc)
    
    # NOTE: Production orders are NO LONGER auto-created here
    # They must be created manually after inventory approval
    # Remove the auto-creation code to enforce approval workflow
    
    return so_obj

@api_router.get("/sales-orders", response_model=List[SalesOrder])
async def get_sales_orders():
    orders = await db.sales_orders.find({}, {"_id": 0}).sort("so_date", -1).to_list(1000)
    for o in orders:
        if isinstance(o.get('so_date'), str):
            o['so_date'] = datetime.fromisoformat(o['so_date'])
        if o.get('expected_delivery_date') and isinstance(o['expected_delivery_date'], str):
            o['expected_delivery_date'] = datetime.fromisoformat(o['expected_delivery_date'])
        if isinstance(o.get('created_at'), str):
            o['created_at'] = datetime.fromisoformat(o['created_at'])
    return orders

@api_router.get("/sales-orders/{so_id}", response_model=SalesOrder)
async def get_sales_order(so_id: str):
    o = await db.sales_orders.find_one({"id": so_id}, {"_id": 0})
    if not o:
        raise HTTPException(status_code=404, detail="Sales order not found")
    if isinstance(o.get('so_date'), str):
        o['so_date'] = datetime.fromisoformat(o['so_date'])
    if o.get('expected_delivery_date') and isinstance(o['expected_delivery_date'], str):
        o['expected_delivery_date'] = datetime.fromisoformat(o['expected_delivery_date'])
    if isinstance(o.get('created_at'), str):
        o['created_at'] = datetime.fromisoformat(o['created_at'])
    return o

@api_router.put("/sales-orders/{so_id}/status")
async def update_sales_order_status(so_id: str, status_data: dict):
    status = status_data.get("so_status")
    if status not in ["draft", "confirmed", "invoiced", "cancelled"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    await db.sales_orders.update_one({"id": so_id}, {"$set": {"so_status": status}})
    return {"message": "Status updated"}

@api_router.delete("/sales-orders/{so_id}")
async def delete_sales_order(so_id: str):
    """Delete a sales order"""
    result = await db.sales_orders.delete_one({"id": so_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sales order not found")
    return {"message": "Sales order deleted successfully"}


# ========== SALES ORDER APPROVAL ROUTES (Inventory Employees) ==========

@api_router.post("/sales-orders/{so_id}/submit-for-approval")
async def submit_sales_order_for_approval(so_id: str):
    """Submit sales order for inventory approval"""
    sales_order = await db.sales_orders.find_one({"id": so_id}, {"_id": 0})
    if not sales_order:
        raise HTTPException(status_code=404, detail="Sales order not found")
    
    if sales_order.get("so_status") != "draft":
        raise HTTPException(status_code=400, detail="Only draft orders can be submitted for approval")
    
    await db.sales_orders.update_one(
        {"id": so_id},
        {"$set": {
            "so_status": "pending_approval",
            "approval_status": "pending"
        }}
    )
    
    return {"message": "Sales order submitted for approval"}


@api_router.post("/sales-orders/{so_id}/approve")
async def approve_sales_order(so_id: str, approval_data: dict):
    """Inventory employee approves sales order"""
    sales_order = await db.sales_orders.find_one({"id": so_id}, {"_id": 0})
    if not sales_order:
        raise HTTPException(status_code=404, detail="Sales order not found")
    
    if sales_order.get("approval_status") != "pending":
        raise HTTPException(status_code=400, detail="Order is not pending approval")
    
    approved_by = approval_data.get("approved_by", "Inventory Manager")
    
    await db.sales_orders.update_one(
        {"id": so_id},
        {"$set": {
            "so_status": "approved",
            "approval_status": "approved",
            "approved_by": approved_by,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Sales order approved by {approved_by}"}


@api_router.post("/sales-orders/{so_id}/reject")
async def reject_sales_order(so_id: str, rejection_data: dict):
    """Inventory employee rejects sales order"""
    sales_order = await db.sales_orders.find_one({"id": so_id}, {"_id": 0})
    if not sales_order:
        raise HTTPException(status_code=404, detail="Sales order not found")
    
    if sales_order.get("approval_status") != "pending":
        raise HTTPException(status_code=400, detail="Order is not pending approval")
    
    rejection_reason = rejection_data.get("reason", "No reason provided")
    rejected_by = rejection_data.get("rejected_by", "Inventory Manager")
    
    await db.sales_orders.update_one(
        {"id": so_id},
        {"$set": {
            "so_status": "rejected",
            "approval_status": "rejected",
            "rejection_reason": rejection_reason,
            "approved_by": rejected_by,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Sales order rejected: {rejection_reason}"}


@api_router.get("/sales-orders/pending-approval/list")
async def get_pending_approval_sales_orders():
    """Get all sales orders pending inventory approval"""
    orders = await db.sales_orders.find(
        {"approval_status": "pending", "so_status": "pending_approval"},
        {"_id": 0}
    ).to_list(length=None)
    return orders


@api_router.get("/sales-orders/approved/list")
async def get_approved_sales_orders():
    """Get all approved sales orders ready for production"""
    orders = await db.sales_orders.find(
        {"approval_status": "approved"},
        {"_id": 0}
    ).to_list(length=None)
    return orders



# ========== DELIVERY CHALLAN ROUTES ==========

@api_router.post("/delivery-challans", response_model=DeliveryChallan)
async def create_delivery_challan(challan_data: DeliveryChallanCreate):
    challan_number = await generate_challan_number()
    challan_obj = DeliveryChallan(challan_number=challan_number, **challan_data.model_dump())
    doc = challan_obj.model_dump()
    doc['challan_date'] = doc['challan_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.delivery_challans.insert_one(doc)
    return challan_obj

@api_router.get("/delivery-challans", response_model=List[DeliveryChallan])
async def get_delivery_challans():
    challans = await db.delivery_challans.find({}, {"_id": 0}).sort("challan_date", -1).to_list(1000)
    for c in challans:
        if isinstance(c.get('challan_date'), str):
            c['challan_date'] = datetime.fromisoformat(c['challan_date'])
        if isinstance(c.get('created_at'), str):
            c['created_at'] = datetime.fromisoformat(c['created_at'])
    return challans

@api_router.get("/delivery-challans/{challan_id}", response_model=DeliveryChallan)
async def get_delivery_challan(challan_id: str):
    c = await db.delivery_challans.find_one({"id": challan_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Delivery challan not found")
    if isinstance(c.get('challan_date'), str):
        c['challan_date'] = datetime.fromisoformat(c['challan_date'])
    if isinstance(c.get('created_at'), str):
        c['created_at'] = datetime.fromisoformat(c['created_at'])
    return c

@api_router.delete("/delivery-challans/{challan_id}")
async def delete_delivery_challan(challan_id: str):
    """Delete a delivery challan"""
    result = await db.delivery_challans.delete_one({"id": challan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Delivery challan not found")


# ========== DISPATCH ROUTES ==========

@api_router.get("/dispatch/ready-orders")
async def get_ready_for_dispatch_orders():
    """Get all sales orders that are ready for dispatch (production completed)"""
    orders = await db.sales_orders.find(
        {"dispatch_status": "ready", "production_status": "completed"},
        {"_id": 0}
    ).to_list(length=None)
    
    # Sort by date
    orders.sort(key=lambda x: x['created_at'], reverse=True)
    
    return orders

@api_router.post("/dispatch/mark-dispatched/{sales_order_id}")
async def mark_order_dispatched(sales_order_id: str):
    """Mark sales order as dispatched"""
    order = await db.sales_orders.find_one({"id": sales_order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Sales order not found")
    
    await db.sales_orders.update_one(
        {"id": sales_order_id},
        {"$set": {"dispatch_status": "dispatched"}}
    )
    
    updated_order = await db.sales_orders.find_one({"id": sales_order_id}, {"_id": 0})
    return updated_order



# ========== EXPENSE ROUTES ==========

@api_router.post("/expenses", response_model=Expense)
async def create_expense(expense_data: ExpenseCreate):
    """Create a new expense entry"""
    # Generate expense number
    exp_count = await db.expenses.count_documents({})
    exp_number = f"EXP-{datetime.now().strftime('%Y%m%d')}-{str(exp_count + 1).zfill(4)}"
    
    expense_obj = Expense(expense_number=exp_number, **expense_data.model_dump())
    doc = expense_obj.model_dump()
    doc['expense_date'] = doc['expense_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.expenses.insert_one(doc)
    return expense_obj


# ========== DAY BOOK ROUTES (Manual Entry) ==========

@api_router.post("/daybook", response_model=DayBookEntry)
async def create_daybook_entry(entry_data: DayBookEntryCreate):
    """Add a manual day book entry"""
    # Parse date
    entry_date = datetime.fromisoformat(entry_data.date.replace('Z', '+00:00'))
    
    # Get previous balance
    previous_entries = await db.daybook.find({}, {"_id": 0}).sort("date", -1).to_list(length=1)
    previous_balance = previous_entries[0]['balance'] if previous_entries else 0.0
    
    # Calculate new balance: previous + credit - debit (bank statement format)
    new_balance = previous_balance + entry_data.credit - entry_data.debit
    
    # Create entry
    entry_obj = DayBookEntry(
        date=entry_date,
        description=entry_data.description,
        purpose=entry_data.purpose,
        debit=entry_data.debit,
        credit=entry_data.credit,
        balance=new_balance
    )
    
    doc = entry_obj.model_dump()
    doc['date'] = doc['date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.daybook.insert_one(doc)
    return entry_obj

@api_router.get("/daybook", response_model=List[DayBookEntry])
async def get_daybook_entries():
    """Get all day book entries"""
    entries = await db.daybook.find({}, {"_id": 0}).to_list(length=None)
    entries.sort(key=lambda x: x['date'])
    return entries

@api_router.get("/daybook/export-excel")
async def export_daybook_excel():
    """Export Day Book to Excel"""
    entries = await db.daybook.find({}, {"_id": 0}).to_list(length=None)
    entries.sort(key=lambda x: x['date'])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Day Book"
    
    # Styling
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font = Font(bold=True)
    green_font = Font(bold=True, color="008000")
    red_font = Font(bold=True, color="FF0000")
    blue_font = Font(bold=True, color="0000FF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = "Day Book - All Transactions"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    row = 3
    headers = ['Date', 'Description', 'Purpose', 'Credit (In)', 'Debit (Out)', 'Balance', 'Created At']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row = 4
    total_debit = 0
    total_credit = 0
    
    for entry in entries:
        ws.cell(row=row, column=1, value=entry.get('date', ''))
        ws.cell(row=row, column=2, value=entry.get('description', ''))
        ws.cell(row=row, column=3, value=entry.get('purpose', ''))
        
        credit_cell = ws.cell(row=row, column=4, value=entry.get('credit', 0))
        credit_cell.number_format = '₹#,##0.00'
        if entry.get('credit', 0) > 0:
            credit_cell.font = green_font  # Credit = Money IN = Green
        
        debit_cell = ws.cell(row=row, column=5, value=entry.get('debit', 0))
        debit_cell.number_format = '₹#,##0.00'
        if entry.get('debit', 0) > 0:
            debit_cell.font = red_font  # Debit = Money OUT = Red
        
        balance_cell = ws.cell(row=row, column=6, value=entry.get('balance', 0))
        balance_cell.number_format = '₹#,##0.00'
        balance_cell.font = blue_font
        
        ws.cell(row=row, column=7, value=entry.get('created_at', ''))
        
        # Apply borders
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
        
        total_credit += entry.get('credit', 0)
        total_debit += entry.get('debit', 0)
        row += 1
    
    # Summary row
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    summary_cell = ws.cell(row=row, column=1, value="TOTAL")
    summary_cell.font = Font(bold=True, size=12)
    summary_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    total_credit_cell = ws.cell(row=row, column=4, value=total_credit)
    total_credit_cell.number_format = '₹#,##0.00'
    total_credit_cell.font = Font(bold=True, color="008000", size=12)  # Green for IN
    total_credit_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    total_debit_cell = ws.cell(row=row, column=5, value=total_debit)
    total_debit_cell.number_format = '₹#,##0.00'
    total_debit_cell.font = Font(bold=True, color="FF0000", size=12)  # Red for OUT
    total_debit_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    final_balance = total_credit - total_debit  # Credit - Debit
    final_balance_cell = ws.cell(row=row, column=6, value=final_balance)
    final_balance_cell.number_format = '₹#,##0.00'
    final_balance_cell.font = Font(bold=True, color="0000FF", size=12)
    final_balance_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws.cell(row=row, column=7, value="")
    
    # Apply borders to summary
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    filename = f"daybook_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/daybook/export-pdf")
async def export_daybook_pdf():
    """Export Day Book to PDF"""
    entries = await db.daybook.find({}, {"_id": 0}).to_list(length=None)
    entries.sort(key=lambda x: x['date'])
    
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18,
                                 textColor=colors.HexColor('#1a1a1a'), spaceAfter=15, alignment=TA_CENTER)
    
    # Title
    title = Paragraph("<b>Day Book - All Transactions</b>", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    table_data = [['Date', 'Description', 'Purpose', 'Credit (In)', 'Debit (Out)', 'Balance']]
    
    total_credit = 0
    total_debit = 0
    
    for entry in entries:
        date_str = entry.get('date', '')
        if isinstance(date_str, str) and 'T' in date_str:
            date_str = date_str.split('T')[0]
        
        credit_str = f"₹{entry.get('credit', 0):,.2f}" if entry.get('credit', 0) > 0 else "-"
        debit_str = f"₹{entry.get('debit', 0):,.2f}" if entry.get('debit', 0) > 0 else "-"
        balance_str = f"₹{entry.get('balance', 0):,.2f}"
        
        table_data.append([
            date_str,
            entry.get('description', '')[:30],
            entry.get('purpose', '')[:20] or '-',
            credit_str,
            debit_str,
            balance_str
        ])
        
        total_credit += entry.get('credit', 0)
        total_debit += entry.get('debit', 0)
    
    # Add summary row
    final_balance = total_credit - total_debit  # Credit - Debit
    table_data.append([
        '',
        'TOTAL',
        '',
        f"₹{total_credit:,.2f}",
        f"₹{total_debit:,.2f}",
        f"₹{final_balance:,.2f}"
    ])
    
    # Create table
    col_widths = [0.8*inch, 1.8*inch, 1.3*inch, 1*inch, 1*inch, 1*inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        
        # Summary row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E7E6E6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('TEXTCOLOR', (3, -1), (3, -1), colors.HexColor('#008000')),
        ('TEXTCOLOR', (4, -1), (4, -1), colors.HexColor('#FF0000')),
        ('TEXTCOLOR', (5, -1), (5, -1), colors.HexColor('#0000FF')),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    filename = f"daybook_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.delete("/daybook/{entry_id}")
async def delete_daybook_entry(entry_id: str):
    """Delete a day book entry and recalculate balances"""
    # Delete the entry
    result = await db.daybook.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Recalculate all balances
    all_entries = await db.daybook.find({}, {"_id": 0}).to_list(length=None)
    all_entries.sort(key=lambda x: x['date'])
    
    running_balance = 0.0
    for entry in all_entries:
        running_balance += entry['credit'] - entry['debit']  # Bank statement format
        await db.daybook.update_one(
            {"id": entry['id']},
            {"$set": {"balance": running_balance}}
        )
    
    return {"message": "Entry deleted and balances recalculated"}


@api_router.put("/daybook/{entry_id}", response_model=DayBookEntry)
async def update_daybook_entry(entry_id: str, entry_data: DayBookEntryCreate):
    """Update a day book entry and recalculate all balances"""
    # Check if entry exists
    existing_entry = await db.daybook.find_one({"id": entry_id}, {"_id": 0})
    if not existing_entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Parse date
    entry_date = datetime.fromisoformat(entry_data.date.replace('Z', '+00:00'))
    
    # Update the entry (without balance for now)
    update_data = {
        "date": entry_date.isoformat(),
        "description": entry_data.description,
        "purpose": entry_data.purpose,
        "debit": entry_data.debit,
        "credit": entry_data.credit
    }
    
    await db.daybook.update_one(
        {"id": entry_id},
        {"$set": update_data}
    )
    
    # Recalculate all balances in chronological order
    all_entries = await db.daybook.find({}, {"_id": 0}).to_list(length=None)
    all_entries.sort(key=lambda x: x['date'])
    
    running_balance = 0.0
    for entry in all_entries:
        running_balance += entry['credit'] - entry['debit']  # Bank statement format
        await db.daybook.update_one(
            {"id": entry['id']},
            {"$set": {"balance": running_balance}}
        )
    
    # Get updated entry
    updated_entry = await db.daybook.find_one({"id": entry_id}, {"_id": 0})
    if isinstance(updated_entry.get('date'), str):
        updated_entry['date'] = datetime.fromisoformat(updated_entry['date'])
    if isinstance(updated_entry.get('created_at'), str):
        updated_entry['created_at'] = datetime.fromisoformat(updated_entry['created_at'])
    
    return updated_entry




@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses():
    """Get all expenses"""
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(length=None)
    expenses.sort(key=lambda x: x['expense_date'], reverse=True)
    return expenses

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str):
    """Delete an expense"""
    result = await db.expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense deleted successfully"}


# ========== FINANCE / DAY BOOK ROUTES ==========

@api_router.get("/finance/daybook")
async def get_daybook(start_date: str = None, end_date: str = None):
    """
    Get day book - all financial transactions for a date range
    Returns all money in/out transactions with running balance
    """
    # Parse dates
    if start_date:
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
    else:
        start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    if end_date:
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    else:
        end = datetime.now(timezone.utc).replace(hour=23, minute=59, second=59, microsecond=999999)
    
    transactions = []
    
    # 1. Sales Invoices (Money In)
    invoices = await db.invoices.find({
        "invoice_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    
    for inv in invoices:
        transactions.append({
            "date": inv['invoice_date'],
            "type": "sale",
            "category": "Sales Invoice",
            "reference": inv['invoice_number'],
            "description": f"Invoice to {inv['customer_name']}",
            "customer_vendor": inv['customer_name'],
            "debit": inv['grand_total'],  # Money In
            "credit": 0,
            "payment_status": inv.get('payment_status', 'unpaid')
        })
    
    # 2. Payments Received (Money In)
    payments_received = await db.payments_received.find({
        "payment_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    
    for pmt in payments_received:
        transactions.append({
            "date": pmt['payment_date'],
            "type": "payment_in",
            "category": "Payment Received",
            "reference": pmt['payment_number'],
            "description": f"Payment from {pmt['customer_name']}",
            "customer_vendor": pmt['customer_name'],
            "debit": pmt['payment_amount'],  # Money In
            "credit": 0,
            "payment_status": "paid"
        })
    
    # 3. Credit Notes (Money Out - Refunds)
    credit_notes = await db.credit_notes.find({
        "credit_note_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    
    for cn in credit_notes:
        transactions.append({
            "date": cn['credit_note_date'],
            "type": "refund",
            "category": "Credit Note",
            "reference": cn['credit_note_number'],
            "description": f"Credit note to {cn['customer_name']} - {cn.get('reason', '')}",
            "customer_vendor": cn['customer_name'],
            "debit": 0,
            "credit": cn['credit_amount'],  # Money Out
            "payment_status": "refunded"
        })
    
    # 4. Purchase Invoices (Money Out)
    purchases = await db.purchase_invoices.find({
        "invoice_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    
    for pur in purchases:
        transactions.append({
            "date": pur['invoice_date'],
            "type": "purchase",
            "category": "Purchase Invoice",
            "reference": pur['invoice_number'],
            "description": f"Purchase from {pur['vendor_name']}",
            "customer_vendor": pur['vendor_name'],
            "debit": 0,
            "credit": pur['grand_total'],  # Money Out
            "payment_status": pur.get('payment_status', 'unpaid')
        })
    
    # 5. Payments Made (Money Out)
    payments_made = await db.payments_made.find({
        "payment_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    
    for pmt in payments_made:
        transactions.append({
            "date": pmt['payment_date'],
            "type": "payment_out",
            "category": "Payment Made",
            "reference": pmt['payment_number'],
            "description": f"Payment to {pmt['vendor_name']}",
            "customer_vendor": pmt['vendor_name'],
            "debit": 0,
            "credit": pmt['payment_amount'],  # Money Out
            "payment_status": "paid"
        })
    
    # 6. Expenses (Money Out)
    expenses = await db.expenses.find({
        "expense_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    
    for exp in expenses:
        transactions.append({
            "date": exp['expense_date'],
            "type": "expense",
            "category": f"Expense - {exp['category']}",
            "reference": exp['expense_number'],
            "description": exp['description'],
            "customer_vendor": exp.get('vendor_name', ''),
            "debit": 0,
            "credit": exp['amount'],  # Money Out
            "payment_status": "paid"
        })
    
    # Sort by date
    transactions.sort(key=lambda x: x['date'])
    
    # Calculate running balance
    running_balance = 0
    for txn in transactions:
        running_balance += txn['debit'] - txn['credit']
        txn['balance'] = running_balance
    
    return {
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "transactions": transactions,
        "opening_balance": 0,  # Can be fetched from previous period
        "closing_balance": running_balance
    }


@api_router.get("/finance/summary")
async def get_finance_summary(start_date: str = None, end_date: str = None):
    """Get financial summary for a date range"""
    # Parse dates
    if start_date:
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
    else:
        start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    if end_date:
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    else:
        end = datetime.now(timezone.utc).replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Calculate totals
    invoices = await db.invoices.find({
        "invoice_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    total_sales = sum(inv['grand_total'] for inv in invoices)
    
    payments_received = await db.payments_received.find({
        "payment_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    total_payments_received = sum(pmt['payment_amount'] for pmt in payments_received)
    
    credit_notes = await db.credit_notes.find({
        "credit_note_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    total_refunds = sum(cn['credit_amount'] for cn in credit_notes)
    
    purchases = await db.purchase_invoices.find({
        "invoice_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    total_purchases = sum(pur['grand_total'] for pur in purchases)
    
    payments_made = await db.payments_made.find({
        "payment_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    total_payments_made = sum(pmt['payment_amount'] for pmt in payments_made)
    
    expenses = await db.expenses.find({
        "expense_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}
    }, {"_id": 0}).to_list(length=None)
    total_expenses = sum(exp['amount'] for exp in expenses)
    
    # Calculate net cash flow
    total_money_in = total_sales + total_payments_received
    total_money_out = total_refunds + total_purchases + total_payments_made + total_expenses
    net_cash_flow = total_money_in - total_money_out
    
    return {
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "total_sales": total_sales,
        "total_payments_received": total_payments_received,
        "total_refunds": total_refunds,
        "total_purchases": total_purchases,
        "total_payments_made": total_payments_made,
        "total_expenses": total_expenses,
        "total_money_in": total_money_in,
        "total_money_out": total_money_out,
        "net_cash_flow": net_cash_flow,
        "invoice_count": len(invoices),
        "payment_received_count": len(payments_received),
        "purchase_count": len(purchases),
        "expense_count": len(expenses)
    }


@api_router.post("/delivery-challans/{challan_id}/update-status")
async def update_delivery_challan_status(challan_id: str, status: str):
    """Update delivery challan dispatch status"""
    valid_statuses = ["pending", "dispatched", "in_transit", "delivered"]
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of {valid_statuses}")
    
    challan = await db.delivery_challans.find_one({"id": challan_id}, {"_id": 0})
    if not challan:
        raise HTTPException(status_code=404, detail="Delivery challan not found")
    
    await db.delivery_challans.update_one(
        {"id": challan_id},
        {"$set": {"dispatch_status": status}}
    )
    
    # If delivered, update sales order status
    if status == "delivered" and challan.get('sales_order_id'):
        await db.sales_orders.update_one(
            {"id": challan['sales_order_id']},
            {"$set": {"dispatch_status": "delivered"}}
        )
    
    updated_challan = await db.delivery_challans.find_one({"id": challan_id}, {"_id": 0})
    return updated_challan


    return {"message": "Delivery challan deleted successfully"}


# ========== CREDIT NOTE ROUTES ==========

@api_router.post("/credit-notes", response_model=CreditNote)
async def create_credit_note(cn_data: CreditNoteCreate):
    cn_number = await generate_credit_note_number()
    totals = await calculate_document_totals(cn_data.items, "", "percentage", 0.0)
    cn_obj = CreditNote(
        credit_note_number=cn_number, **cn_data.model_dump(),
        subtotal=totals['subtotal'], total_discount=totals['total_discount'],
        taxable_amount=totals['taxable_amount'], is_interstate=totals['is_interstate'],
        cgst_amount=totals['cgst_amount'], sgst_amount=totals['sgst_amount'],
        igst_amount=totals['igst_amount'], total_gst=totals['total_gst'],
        credit_amount=totals['grand_total'], stock_restored=True
    )
    await restore_stock_on_return(cn_data.items)
    doc = cn_obj.model_dump()
    doc['credit_note_date'] = doc['credit_note_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.credit_notes.insert_one(doc)
    return cn_obj

@api_router.get("/credit-notes", response_model=List[CreditNote])
async def get_credit_notes():
    cns = await db.credit_notes.find({}, {"_id": 0}).sort("credit_note_date", -1).to_list(1000)
    for cn in cns:
        if isinstance(cn.get('credit_note_date'), str):
            cn['credit_note_date'] = datetime.fromisoformat(cn['credit_note_date'])
        if isinstance(cn.get('created_at'), str):
            cn['created_at'] = datetime.fromisoformat(cn['created_at'])
    return cns

@api_router.get("/credit-notes/{cn_id}", response_model=CreditNote)
async def get_credit_note(cn_id: str):
    cn = await db.credit_notes.find_one({"id": cn_id}, {"_id": 0})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    if isinstance(cn.get('credit_note_date'), str):
        cn['credit_note_date'] = datetime.fromisoformat(cn['credit_note_date'])
    if isinstance(cn.get('created_at'), str):
        cn['created_at'] = datetime.fromisoformat(cn['created_at'])
    return cn

@api_router.delete("/credit-notes/{cn_id}")
async def delete_credit_note(cn_id: str):
    """Delete a credit note and reverse stock restoration"""
    # Get the credit note first
    cn = await db.credit_notes.find_one({"id": cn_id})
    if not cn:
        raise HTTPException(status_code=404, detail="Credit note not found")
    
    # If stock was restored, reverse it
    if cn.get('stock_restored'):
        for item in cn.get('items', []):
            await db.products.update_one(
                {"id": item['product_id']},
                {"$inc": {"stock_quantity": -item['quantity']}}
            )
    
    result = await db.credit_notes.delete_one({"id": cn_id})
    return {"message": "Credit note deleted successfully"}


# ========== REPORTS ROUTES ==========

@api_router.get("/reports/sales")
async def get_sales_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    customer_id: Optional[str] = None,
    product_id: Optional[str] = None
):
    query = {}
    if start_date and end_date:
        query['invoice_date'] = {"$gte": start_date, "$lte": end_date}
    if customer_id:
        query['customer_id'] = customer_id
    
    invoices = await db.invoices.find(query, {"_id": 0}).to_list(10000)
    
    if product_id:
        invoices = [inv for inv in invoices if any(item['product_id'] == product_id for item in inv.get('items', []))]
    
    total_sales = sum(inv.get('grand_total', 0) for inv in invoices)
    
    product_sales = {}
    for inv in invoices:
        for item in inv.get('items', []):
            pid = item['product_id']
            if pid not in product_sales:
                product_sales[pid] = {"product_name": item['product_name'], "quantity_sold": 0, "total_revenue": 0}
            item_total = item['quantity'] * item['price']
            item_discount = item_total * (item.get('discount_percent', 0) / 100)
            item_taxable = item_total - item_discount
            item_gst = item_taxable * (item.get('gst_rate', 0) / 100)
            product_sales[pid]["quantity_sold"] += item['quantity']
            product_sales[pid]["total_revenue"] += item_taxable + item_gst
    
    return {
        "summary": {"total_invoices": len(invoices), "total_sales": round(total_sales, 2)},
        "product_wise": list(product_sales.values()),
        "invoices": invoices
    }

@api_router.get("/reports/customer-ledger/{customer_id}")
async def get_customer_ledger(customer_id: str):
    """Get customer ledger with invoices, credit notes, payments, and journal entries"""
    # Get all transactions
    invoices = await db.invoices.find({"customer_id": customer_id}, {"_id": 0}).sort("invoice_date", 1).to_list(10000)
    credit_notes = await db.credit_notes.find({"customer_id": customer_id}, {"_id": 0}).sort("credit_note_date", 1).to_list(10000)
    payments = await db.payments.find({"partner_id": customer_id, "payment_type": "receive"}, {"_id": 0}).sort("payment_date", 1).to_list(10000)
    journal_entries = await db.journal_entries.find({"customer_id": customer_id}, {"_id": 0}).sort("entry_date", 1).to_list(10000)
    
    # Calculate totals
    total_invoiced = sum(inv.get('grand_total', 0) for inv in invoices if inv.get('invoice_status') != 'cancelled')
    total_credited = sum(cn.get('credit_amount', 0) for cn in credit_notes)
    total_paid = sum(p.get('payment_amount', 0) for p in payments)
    
    # Calculate journal entry impact
    # Opening balance: positive = customer owes (Debit), negative = you owe customer (Credit)
    # Freight paid by customer = Credit (reduces balance)
    # Discounts given = Credit (reduces balance)
    # Other charges = Debit (increases balance)
    journal_debits = 0
    journal_credits = 0
    
    for je in journal_entries:
        amount = je.get('amount', 0)
        entry_type = je.get('entry_type', '')
        
        if entry_type == 'opening_balance':
            # Opening balance: positive = Debit, negative = Credit
            if amount >= 0:
                journal_debits += amount
            else:
                journal_credits += abs(amount)
        elif entry_type == 'freight':
            # Freight paid by customer on behalf of company = Credit
            journal_credits += abs(amount)
        elif entry_type == 'discount' or amount < 0:
            # Discounts = Credit
            journal_credits += abs(amount)
        else:
            # Other charges = Debit
            journal_debits += abs(amount)
    
    net_balance = total_invoiced + journal_debits - total_credited - journal_credits - total_paid
    
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    
    return {
        "customer": customer,
        "summary": {
            "total_invoiced": round(total_invoiced, 2),
            "total_credited": round(total_credited, 2),
            "total_paid": round(total_paid, 2),
            "journal_debits": round(journal_debits, 2),
            "journal_credits": round(journal_credits, 2),
            "net_balance": round(net_balance, 2)
        },
        "invoices": invoices,
        "credit_notes": credit_notes,
        "payments": payments,
        "journal_entries": journal_entries
    }


# ========== PAYMENT ROUTES ==========

@api_router.post("/payments", response_model=Payment)
async def create_payment(payment_data: PaymentCreate):
    """Create a payment (received or made) and update invoice statuses"""
    # Generate payment number
    payment_number = await generate_payment_number(payment_data.payment_type)
    
    # Calculate unallocated amount
    total_allocated = sum(allocation.allocated_amount for allocation in payment_data.allocations)
    unallocated_amount = payment_data.payment_amount - total_allocated
    
    payment_obj = Payment(
        payment_number=payment_number,
        **payment_data.model_dump(),
        unallocated_amount=unallocated_amount
    )
    
    doc = payment_obj.model_dump()
    doc['payment_date'] = doc['payment_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.payments.insert_one(doc)
    
    # Update invoice payment statuses
    for allocation in payment_data.allocations:
        await update_invoice_payment_status_from_payment(
            allocation.invoice_id,
            allocation.invoice_type
        )
    
    return payment_obj

@api_router.get("/payments", response_model=List[Payment])
async def get_payments(
    payment_type: Optional[str] = None,
    partner_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Get all payments with optional filters"""
    query = {}
    
    if payment_type:
        query['payment_type'] = payment_type
    
    if partner_id:
        query['partner_id'] = partner_id
    
    if start_date and end_date:
        query['payment_date'] = {"$gte": start_date, "$lte": end_date}
    
    payments = await db.payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(10000)
    
    for payment in payments:
        if isinstance(payment.get('payment_date'), str):
            payment['payment_date'] = datetime.fromisoformat(payment['payment_date'])
        if isinstance(payment.get('created_at'), str):
            payment['created_at'] = datetime.fromisoformat(payment['created_at'])
    
    return payments

@api_router.get("/payments/{payment_id}", response_model=Payment)
async def get_payment(payment_id: str):
    """Get single payment by ID"""
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    if isinstance(payment.get('payment_date'), str):
        payment['payment_date'] = datetime.fromisoformat(payment['payment_date'])
    if isinstance(payment.get('created_at'), str):
        payment['created_at'] = datetime.fromisoformat(payment['created_at'])
    
    return payment

@api_router.get("/payments/invoice/{invoice_id}")
async def get_payments_for_invoice(invoice_id: str, invoice_type: str = "sales_invoice"):
    """Get all payments allocated to a specific invoice"""
    all_payments = await db.payments.find({}, {"_id": 0}).to_list(10000)
    
    invoice_payments = []
    for payment in all_payments:
        for allocation in payment.get('allocations', []):
            if allocation['invoice_id'] == invoice_id and allocation['invoice_type'] == invoice_type:
                if isinstance(payment.get('payment_date'), str):
                    payment['payment_date'] = datetime.fromisoformat(payment['payment_date'])
                if isinstance(payment.get('created_at'), str):
                    payment['created_at'] = datetime.fromisoformat(payment['created_at'])
                invoice_payments.append({
                    "payment": payment,
                    "allocated_amount": allocation['allocated_amount']
                })
                break
    
    return invoice_payments

@api_router.put("/payments/{payment_id}/status")
async def update_payment_status(payment_id: str, status_data: dict):
    """Update payment status (posted, reconciled)"""
    status = status_data.get("status")
    if status not in ["draft", "posted", "reconciled"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.payments.update_one(
        {"id": payment_id},
        {"$set": {"status": status}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    return {"message": "Payment status updated successfully"}

@api_router.delete("/payments/{payment_id}")
async def delete_payment(payment_id: str):
    """Delete a payment and update invoice payment statuses"""
    # Get the payment first
    payment = await db.payments.find_one({"id": payment_id})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Update invoice payment statuses for all allocations
    for allocation in payment.get('allocations', []):
        await update_invoice_payment_status_from_payment(
            allocation['invoice_id'],
            allocation['invoice_type']
        )
    
    result = await db.payments.delete_one({"id": payment_id})
    return {"message": "Payment deleted successfully"}

# ========== JOURNAL ENTRY ROUTES ==========

@api_router.post("/journal-entries", response_model=JournalEntry)
async def create_journal_entry(entry_data: JournalEntryCreate):
    """Create a new journal entry"""
    entry_number = await generate_journal_entry_number()
    entry_obj = JournalEntry(entry_number=entry_number, **entry_data.model_dump())
    doc = entry_obj.model_dump()
    doc['entry_date'] = doc['entry_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.journal_entries.insert_one(doc)
    return entry_obj

@api_router.get("/journal-entries", response_model=List[JournalEntry])
async def get_journal_entries(
    customer_id: Optional[str] = None,
    entry_type: Optional[str] = None
):
    """Get all journal entries with optional filters"""
    query = {}
    if customer_id:
        query['customer_id'] = customer_id
    if entry_type:
        query['entry_type'] = entry_type
    
    entries = await db.journal_entries.find(query, {"_id": 0}).sort("entry_date", -1).to_list(1000)
    for entry in entries:
        if isinstance(entry.get('entry_date'), str):
            entry['entry_date'] = datetime.fromisoformat(entry['entry_date'])
        if isinstance(entry.get('created_at'), str):
            entry['created_at'] = datetime.fromisoformat(entry['created_at'])
    return entries

@api_router.get("/journal-entries/{entry_id}", response_model=JournalEntry)
async def get_journal_entry(entry_id: str):
    """Get a specific journal entry"""
    entry = await db.journal_entries.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    if isinstance(entry.get('entry_date'), str):
        entry['entry_date'] = datetime.fromisoformat(entry['entry_date'])
    if isinstance(entry.get('created_at'), str):
        entry['created_at'] = datetime.fromisoformat(entry['created_at'])
    return entry

@api_router.put("/journal-entries/{entry_id}", response_model=JournalEntry)
async def update_journal_entry(entry_id: str, entry_data: JournalEntryCreate):
    """Update a journal entry"""
    existing = await db.journal_entries.find_one({"id": entry_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    
    update_data = entry_data.model_dump()
    await db.journal_entries.update_one(
        {"id": entry_id},
        {"$set": update_data}
    )
    
    updated_entry = await db.journal_entries.find_one({"id": entry_id}, {"_id": 0})
    if isinstance(updated_entry.get('entry_date'), str):
        updated_entry['entry_date'] = datetime.fromisoformat(updated_entry['entry_date'])
    if isinstance(updated_entry.get('created_at'), str):
        updated_entry['created_at'] = datetime.fromisoformat(updated_entry['created_at'])
    return updated_entry

@api_router.delete("/journal-entries/{entry_id}")
async def delete_journal_entry(entry_id: str):
    """Delete a journal entry"""
    result = await db.journal_entries.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    return {"message": "Journal entry deleted successfully"}

@api_router.get("/reports/payment-summary")
async def get_payment_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Get payment summary report"""
    query = {}
    
    if start_date and end_date:
        query['payment_date'] = {"$gte": start_date, "$lte": end_date}
    
    payments = await db.payments.find(query, {"_id": 0}).to_list(10000)
    
    # Calculate summaries
    total_received = sum(p['payment_amount'] for p in payments if p['payment_type'] == 'receive')
    total_paid = sum(p['payment_amount'] for p in payments if p['payment_type'] == 'pay')
    
    # Payment method breakdown
    method_breakdown = {}
    for payment in payments:
        method = payment.get('payment_method', 'unknown')
        if method not in method_breakdown:
            method_breakdown[method] = {"received": 0, "paid": 0, "count": 0}
        
        if payment['payment_type'] == 'receive':
            method_breakdown[method]["received"] += payment['payment_amount']
        else:
            method_breakdown[method]["paid"] += payment['payment_amount']
        method_breakdown[method]["count"] += 1
    
    return {
        "summary": {
            "total_received": round(total_received, 2),
            "total_paid": round(total_paid, 2),
            "net_cashflow": round(total_received - total_paid, 2),
            "total_transactions": len(payments)
        },
        "payment_methods": method_breakdown,
        "payments": payments
    }


@api_router.get("/reports/outstanding-invoices")
async def get_outstanding_invoices(invoice_type: str = "sales"):
    """Get all unpaid and partially paid invoices"""
    if invoice_type == "sales":
        collection = db.invoices
        doc_type = "sales_invoice"
    else:
        collection = db.purchase_invoices
        doc_type = "purchase_invoice"
    
    # Get invoices that are not fully paid
    invoices = await collection.find(
        {"payment_status": {"$in": ["unpaid", "partial"]}},
        {"_id": 0}
    ).sort("invoice_date", -1).to_list(10000)
    
    # Calculate outstanding amount for each invoice
    all_payments = await db.payments.find({}, {"_id": 0}).to_list(10000)
    
    result = []
    for invoice in invoices:
        # Calculate total paid for this invoice
        total_paid = 0
        for payment in all_payments:
            for allocation in payment.get('allocations', []):
                if allocation['invoice_id'] == invoice['id']:
                    total_paid += allocation['allocated_amount']
        
        outstanding = invoice.get('grand_total', 0) - total_paid
        
        result.append({
            "invoice_id": invoice['id'],
            "invoice_number": invoice['invoice_number'],
            "invoice_date": invoice.get('invoice_date'),
            "partner_name": invoice.get('customer_name') if invoice_type == "sales" else invoice.get('supplier_name'),
            "partner_id": invoice.get('customer_id') if invoice_type == "sales" else invoice.get('supplier_id'),
            "grand_total": invoice.get('grand_total', 0),
            "total_paid": round(total_paid, 2),
            "outstanding_amount": round(outstanding, 2),
            "payment_status": invoice.get('payment_status'),
            "days_overdue": (datetime.now(timezone.utc) - datetime.fromisoformat(invoice['invoice_date']) if isinstance(invoice.get('invoice_date'), str) else datetime.now(timezone.utc) - invoice['invoice_date']).days if invoice.get('invoice_date') else 0
        })
    
    return {
        "total_outstanding": round(sum(inv['outstanding_amount'] for inv in result), 2),
        "invoice_count": len(result),
        "invoices": result
    }


@api_router.get("/dashboard/stats")
async def get_dashboard_stats():
    total_invoices = await db.invoices.count_documents({})
    
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(10000)
    total_revenue = sum(inv.get('grand_total', 0) for inv in invoices)
    
    unpaid_invoices = [inv for inv in invoices if inv.get('payment_status') == 'unpaid']
    pending_amount = sum(inv.get('grand_total', 0) for inv in unpaid_invoices)
    
    total_products = await db.products.count_documents({})
    total_customers = await db.customers.count_documents({})
    
    return {
        "total_invoices": total_invoices,
        "total_revenue": round(total_revenue, 2),
        "pending_amount": round(pending_amount, 2),
        "total_products": total_products,
        "total_customers": total_customers,
        "paid_invoices": len([inv for inv in invoices if inv.get('payment_status') == 'paid']),
        "unpaid_invoices": len(unpaid_invoices)
    }




# ========== BOM ROUTES ==========

@api_router.post("/bom", response_model=BOM)
async def create_bom(bom: BOMCreate):
    """Create a new BOM for a product"""
    bom_data = bom.model_dump()
    bom_data['id'] = str(uuid4())
    bom_data['created_at'] = datetime.now(timezone.utc).isoformat()
    bom_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.boms.insert_one(bom_data)
    return bom_data

@api_router.get("/bom", response_model=List[BOM])
async def get_all_boms():
    """Get all BOMs"""
    boms = await db.boms.find({}, {"_id": 0}).to_list(1000)
    return boms

@api_router.get("/bom/product/{product_id}", response_model=BOM)
async def get_bom_by_product(product_id: str):
    """Get BOM for a specific product"""
    bom = await db.boms.find_one({"product_id": product_id}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found for this product")
    return bom

@api_router.get("/bom/{bom_id}", response_model=BOM)
async def get_bom(bom_id: str):
    """Get a specific BOM"""
    bom = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    return bom

@api_router.put("/bom/{bom_id}", response_model=BOM)
async def update_bom(bom_id: str, bom_update: BOMUpdate):
    """Update a BOM"""
    update_data = {k: v for k, v in bom_update.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.boms.update_one(
        {"id": bom_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    updated_bom = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    return updated_bom

@api_router.delete("/bom/{bom_id}")
async def delete_bom(bom_id: str):
    """Delete a BOM"""
    result = await db.boms.delete_one({"id": bom_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="BOM not found")
    return {"message": "BOM deleted successfully"}


# ========== PRODUCTION ORDER ROUTES ==========

@api_router.post("/production-orders", response_model=ProductionOrder)
async def create_production_order(order: ProductionOrderCreate):
    """Create a new production order - Only from approved sales orders"""
    order_data = order.model_dump()
    
    # Validate that sales order is approved by inventory
    if order_data.get('sales_order_id'):
        sales_order = await db.sales_orders.find_one(
            {"id": order_data['sales_order_id']},
            {"_id": 0}
        )
        
        if not sales_order:
            raise HTTPException(
                status_code=404,
                detail="Sales order not found"
            )
        
        if sales_order.get('approval_status') != 'approved':
            raise HTTPException(
                status_code=400,
                detail=f"Cannot create production order. Sales order must be approved by inventory. Current status: {sales_order.get('approval_status', 'unknown')}"
            )
    
    order_data['id'] = str(uuid4())
    order_data['order_number'] = await generate_production_order_number()
    order_data['status'] = 'draft'
    order_data['created_at'] = datetime.now(timezone.utc).isoformat()
    order_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    # If BOM is provided, calculate materials required
    if order_data.get('bom_id'):
        bom = await db.boms.find_one({"id": order_data['bom_id']}, {"_id": 0})
        if bom:
            materials_required = []
            for material in bom.get('materials', []):
                materials_required.append({
                    "material_type": material['material_type'],
                    "material_id": material['material_id'],
                    "material_name": material['material_name'],
                    "required_quantity": material['quantity'] * order_data['quantity_to_produce'],
                    "allocated_quantity": 0.0,
                    "unit": material['unit']
                })
            order_data['materials_required'] = materials_required
    
    await db.production_orders.insert_one(order_data)
    return order_data

@api_router.get("/production-orders", response_model=List[ProductionOrder])
async def get_all_production_orders():
    """Get all production orders"""
    orders = await db.production_orders.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return orders


@api_router.get("/production-orders/pending-approval/list")
async def get_pending_production_orders():
    """Get production orders pending approval (draft status)"""
    orders = await db.production_orders.find(
        {"status": "draft"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    return orders

@api_router.get("/production-orders/{order_id}", response_model=ProductionOrder)
async def get_production_order(order_id: str):
    """Get a specific production order"""
    order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Production order not found")
    return order

@api_router.put("/production-orders/{order_id}", response_model=ProductionOrder)
async def update_production_order(order_id: str, order_update: ProductionOrderUpdate):
    """Update production order status"""
    update_data = {k: v for k, v in order_update.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.production_orders.update_one(
        {"id": order_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Production order not found")
    
    updated_order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    return updated_order

@api_router.post("/production-orders/{order_id}/approve")
async def approve_production_order(order_id: str, approval_data: dict):
    """Approve production order - changes status from draft to scheduled"""
    order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    
    if not order:
        raise HTTPException(status_code=404, detail="Production order not found")
    
    if order['status'] != 'draft':
        raise HTTPException(status_code=400, detail="Only draft production orders can be approved")
    
    approved_by = approval_data.get("approved_by", "Production Manager")
    
    # Update production order status to scheduled
    await db.production_orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": "scheduled",
            "approved_by": approved_by,
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated_order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    return {"message": f"Production order approved by {approved_by}", "order": updated_order}


@api_router.post("/production-orders/{order_id}/reject")
async def reject_production_order(order_id: str, rejection_data: dict):
    """Reject production order"""
    order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    
    if not order:
        raise HTTPException(status_code=404, detail="Production order not found")
    
    if order['status'] != 'draft':
        raise HTTPException(status_code=400, detail="Only draft production orders can be rejected")
    
    rejection_reason = rejection_data.get("reason", "No reason provided")
    rejected_by = rejection_data.get("rejected_by", "Production Manager")
    
    await db.production_orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": "cancelled",
            "rejection_reason": rejection_reason,
            "rejected_by": rejected_by,
            "rejected_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Production order rejected: {rejection_reason}"}


@api_router.post("/production-orders/{order_id}/start")
async def start_production(order_id: str):
    """Start production - change status to in_progress (only for approved/scheduled orders)"""
    order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Production order not found")
    
    if order['status'] != 'scheduled':
        raise HTTPException(status_code=400, detail="Production order must be approved (scheduled) before starting")
    
    # Update status and start date
    await db.production_orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": "in_progress",
            "start_date": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    updated_order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    return updated_order

@api_router.post("/production-orders/{order_id}/complete")
async def complete_production(order_id: str):
    """Complete production - deduct materials and add finished goods to stock"""
    order = await db.production_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Production order not found")
    
    if order['status'] != 'in_progress':
        raise HTTPException(status_code=400, detail="Production must be in progress to complete")
    
    # Deduct materials from stock
    for material in order.get('materials_required', []):
        if material['material_type'] == 'raw':
            await db.raw_materials.update_one(
                {"id": material['material_id']},
                {"$inc": {"stock_quantity": -material['required_quantity']}}
            )
        elif material['material_type'] == 'packing':
            await db.packing_materials.update_one(
                {"id": material['material_id']},
                {"$inc": {"stock_quantity": -material['required_quantity']}}
            )
    
    # Add finished goods to product stock
    await db.products.update_one(
        {"id": order['product_id']},
        {"$inc": {"stock_quantity": order['quantity_to_produce']}}
    )
    
    # Update production order status
    await db.production_orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": "completed",
            "completion_date": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update linked sales order production status if exists
    if order.get('sales_order_id'):
        # Check if all production orders for this sales order are completed
        sales_order_id = order['sales_order_id']


# ========== MATERIAL REQUEST ROUTES ==========

@api_router.post("/material-requests", response_model=MaterialRequest)
async def create_material_request(request_data: MaterialRequestCreate):
    """Production manager creates material request for inventory approval"""
    # Get production order details
    production_order = await db.production_orders.find_one(
        {"id": request_data.production_order_id},
        {"_id": 0}
    )
    
    if not production_order:
        raise HTTPException(status_code=404, detail="Production order not found")
    
    # Generate request number
    count = await db.material_requests.count_documents({})
    request_number = f"MR-{datetime.now().strftime('%Y%m%d')}-{str(count + 1).zfill(4)}"
    
    # Create material request
    request_obj = MaterialRequest(
        request_number=request_number,
        production_order_number=production_order['order_number'],
        **request_data.model_dump()
    )
    
    doc = request_obj.model_dump()
    doc['request_date'] = doc['request_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.material_requests.insert_one(doc)
    
    return request_obj


@api_router.get("/material-requests/pending")
async def get_pending_material_requests():
    """Get all pending material requests for inventory approval"""
    requests = await db.material_requests.find(
        {"status": "pending"},
        {"_id": 0}
    ).to_list(length=None)
    return requests


@api_router.post("/material-requests/{request_id}/approve")
async def approve_material_request(request_id: str, approval_data: dict):
    """Inventory manager approves material request"""
    request = await db.material_requests.find_one({"id": request_id}, {"_id": 0})
    
    if not request:
        raise HTTPException(status_code=404, detail="Material request not found")
    
    if request.get('status') != 'pending':
        raise HTTPException(status_code=400, detail="Request is not pending")
    
    approved_by = approval_data.get("approved_by", "Inventory Manager")
    
    # Update material request
    await db.material_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": approved_by,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update production order status to indicate materials are approved
    await db.production_orders.update_one(
        {"id": request['production_order_id']},
        {"$set": {
            "material_status": "approved",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Material request approved by {approved_by}"}


@api_router.post("/material-requests/{request_id}/reject")
async def reject_material_request(request_id: str, rejection_data: dict):
    """Inventory manager rejects material request"""
    request = await db.material_requests.find_one({"id": request_id}, {"_id": 0})
    
    if not request:
        raise HTTPException(status_code=404, detail="Material request not found")
    
    if request.get('status') != 'pending':
        raise HTTPException(status_code=400, detail="Request is not pending")
    
    rejection_reason = rejection_data.get("reason", "No reason provided")
    rejected_by = rejection_data.get("rejected_by", "Inventory Manager")
    
    await db.material_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "rejection_reason": rejection_reason,
            "approved_by": rejected_by,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update production order
    await db.production_orders.update_one(
        {"id": request['production_order_id']},
        {"$set": {
            "material_status": "rejected",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"Material request rejected: {rejection_reason}"}


@api_router.get("/material-requests/by-production-order/{production_order_id}")
async def get_material_requests_for_production_order(production_order_id: str):
    """Get material requests for a specific production order"""
    requests = await db.material_requests.find(
        {"production_order_id": production_order_id},
        {"_id": 0}
    ).to_list(length=None)
    return requests


@api_router.delete("/production-orders/{order_id}")
async def delete_production_order(order_id: str):
    """Delete a production order (only if draft or cancelled)"""


# ========== CUSTOMER LEDGER EXPORT ROUTES ==========

@api_router.get("/reports/customer-ledger/{customer_id}/export-excel")
async def export_customer_ledger_excel(customer_id: str):
    """Export customer ledger - Each item with individual amount and balance"""
    ledger_response = await get_customer_ledger(customer_id)
    customer = ledger_response['customer']
    summary = ledger_response['summary']
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Ledger"
    
    # Styling
    title_font = Font(bold=True, size=16)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font = Font(bold=True)
    red_font = Font(bold=True, color="FF0000")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = f"Customer Ledger - {customer.get('name', 'N/A')}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Customer Details
    row = 3
    ws[f'A{row}'] = "Customer ID:"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = customer.get('id', 'N/A')
    ws.merge_cells(f'B{row}:D{row}')
    row += 1
    ws[f'A{row}'] = "Address:"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = customer.get('address', 'N/A')
    ws.merge_cells(f'B{row}:D{row}')
    row += 1
    ws[f'A{row}'] = "Phone:"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = customer.get('phone', 'N/A')
    row += 1
    ws[f'A{row}'] = "GST:"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = customer.get('gst_number', 'N/A')
    
    # Summary
    row = 8
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    ws[f'A{row}'] = "Total Invoiced:"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = summary.get('total_invoiced', 0)
    ws[f'B{row}'].number_format = '₹#,##0.00'
    row += 1
    ws[f'A{row}'] = "Total Paid:"
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = summary.get('total_paid', 0)
    ws[f'B{row}'].number_format = '₹#,##0.00'
    row += 1
    ws[f'A{row}'] = "Outstanding Balance:"
    ws[f'A{row}'].font = red_font
    ws[f'B{row}'] = summary.get('net_balance', 0)
    ws[f'B{row}'].number_format = '₹#,##0.00'
    ws[f'B{row}'].font = red_font
    
    # Headers
    row = 14
    headers = ['Date', 'Type', 'Reference', 'Product Name', 'Quantity', 'Rate', 'Discount %', 'GST %', 'Tax Amt', 'Debit', 'Credit', 'Balance']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Process transactions
    row = 15
    running_balance = 0
    all_txns = []
    
    for inv in ledger_response['invoices']:
        all_txns.append(('invoice', inv.get('invoice_date'), inv))
    for cn in ledger_response['credit_notes']:
        all_txns.append(('credit_note', cn.get('credit_note_date'), cn))
    for pmt in ledger_response['payments']:
        all_txns.append(('payment', pmt.get('payment_date'), pmt))
    for je in ledger_response['journal_entries']:
        all_txns.append(('journal', je.get('entry_date'), je))
    
    all_txns.sort(key=lambda x: x[1] if x[1] else datetime.min)
    
    for txn_type, txn_date, data in all_txns:
        date_str = txn_date.strftime('%d/%m/%y') if isinstance(txn_date, datetime) else str(txn_date)
        
        if txn_type == 'invoice':
            if data.get('items'):
                # Each item gets its own row with individual amount and balance
                for idx, item in enumerate(data.get('items', [])):
                    item_total = item.get('quantity', 0) * item.get('price', 0)
                    disc_amt = item_total * (item.get('discount_percent', 0) / 100)
                    taxable = item_total - disc_amt
                    tax_amt = taxable * (item.get('gst_rate', 0) / 100)
                    item_final = taxable + tax_amt
                    
                    # Update running balance for each item
                    running_balance += item_final
                    
                    # Date, Type, Reference
                    ws.cell(row, 1).value = date_str if idx == 0 else ''
                    ws.cell(row, 2).value = 'Invoice'
                    ws.cell(row, 3).value = data.get('invoice_number')
                    
                    # Item details
                    ws.cell(row, 4).value = item.get('product_name', '')
                    ws.cell(row, 5).value = f"{item.get('quantity', 0)} {item.get('unit', '')}"
                    ws.cell(row, 6).value = item.get('price', 0)
                    ws.cell(row, 6).number_format = '₹#,##0.00'
                    ws.cell(row, 7).value = f"{item.get('discount_percent', 0)}%"
                    ws.cell(row, 8).value = f"{item.get('gst_rate', 0)}%"
                    ws.cell(row, 9).value = tax_amt
                    ws.cell(row, 9).number_format = '₹#,##0.00'
                    
                    # Individual item amount and balance
                    ws.cell(row, 10).value = item_final
                    ws.cell(row, 10).number_format = '₹#,##0.00'
                    ws.cell(row, 11).value = ''
                    ws.cell(row, 12).value = running_balance
                    ws.cell(row, 12).number_format = '₹#,##0.00'
                    
                    for c in range(1, 13):
                        ws.cell(row, c).border = border
                    row += 1
            else:
                # Invoice without items
                invoice_total = data.get('grand_total', 0)
                running_balance += invoice_total
                
                ws.cell(row, 1).value = date_str
                ws.cell(row, 2).value = 'Invoice'
                ws.cell(row, 3).value = data.get('invoice_number')
                for c in range(4, 10):
                    ws.cell(row, c).value = ''
                ws.cell(row, 10).value = invoice_total
                ws.cell(row, 10).number_format = '₹#,##0.00'
                ws.cell(row, 11).value = ''
                ws.cell(row, 12).value = running_balance
                ws.cell(row, 12).number_format = '₹#,##0.00'
                
                for c in range(1, 13):
                    ws.cell(row, c).border = border
                row += 1
        
        elif txn_type == 'credit_note':
            if data.get('items'):
                for idx, item in enumerate(data.get('items', [])):
                    item_total = item.get('quantity', 0) * item.get('price', 0)
                    disc_amt = item_total * (item.get('discount_percent', 0) / 100)
                    taxable = item_total - disc_amt
                    tax_amt = taxable * (item.get('gst_rate', 0) / 100)
                    item_final = taxable + tax_amt
                    
                    # Update running balance for each item
                    running_balance -= item_final
                    
                    ws.cell(row, 1).value = date_str if idx == 0 else ''
                    ws.cell(row, 2).value = 'Credit Note'
                    ws.cell(row, 3).value = data.get('credit_note_number')
                    ws.cell(row, 4).value = item.get('product_name', '')
                    ws.cell(row, 5).value = f"{item.get('quantity', 0)} {item.get('unit', '')}"
                    ws.cell(row, 6).value = item.get('price', 0)
                    ws.cell(row, 6).number_format = '₹#,##0.00'
                    ws.cell(row, 7).value = f"{item.get('discount_percent', 0)}%"
                    ws.cell(row, 8).value = f"{item.get('gst_rate', 0)}%"
                    ws.cell(row, 9).value = tax_amt
                    ws.cell(row, 9).number_format = '₹#,##0.00'
                    
                    # Individual item amount and balance
                    ws.cell(row, 10).value = ''
                    ws.cell(row, 11).value = item_final
                    ws.cell(row, 11).number_format = '₹#,##0.00'
                    ws.cell(row, 12).value = running_balance
                    ws.cell(row, 12).number_format = '₹#,##0.00'
                    
                    for c in range(1, 13):
                        ws.cell(row, c).border = border
                    row += 1
            else:
                credit_amt = data.get('credit_amount', 0)
                running_balance -= credit_amt
                
                ws.cell(row, 1).value = date_str
                ws.cell(row, 2).value = 'Credit Note'
                ws.cell(row, 3).value = data.get('credit_note_number')
                for c in range(4, 10):
                    ws.cell(row, c).value = ''
                ws.cell(row, 10).value = ''
                ws.cell(row, 11).value = credit_amt
                ws.cell(row, 11).number_format = '₹#,##0.00'
                ws.cell(row, 12).value = running_balance
                ws.cell(row, 12).number_format = '₹#,##0.00'
                
                for c in range(1, 13):
                    ws.cell(row, c).border = border
                row += 1
        
        elif txn_type == 'payment':
            running_balance -= data.get('payment_amount', 0)
            
            ws.cell(row, 1).value = date_str
            ws.cell(row, 2).value = 'Payment'
            ws.cell(row, 3).value = data.get('payment_number')
            for c in range(4, 10):
                ws.cell(row, c).value = ''
            ws.cell(row, 10).value = ''
            ws.cell(row, 11).value = data.get('payment_amount', 0)
            ws.cell(row, 11).number_format = '₹#,##0.00'
            ws.cell(row, 12).value = running_balance
            ws.cell(row, 12).number_format = '₹#,##0.00'
            
            for c in range(1, 13):
                ws.cell(row, c).border = border
            row += 1
        
        elif txn_type == 'journal':
            amount = data.get('amount', 0)
            entry_type = data.get('entry_type', '')
            
            ws.cell(row, 1).value = date_str
            if entry_type == 'opening_balance':
                ws.cell(row, 2).value = 'Opening Balance'
            elif entry_type == 'freight':
                ws.cell(row, 2).value = 'Freight'
            elif entry_type == 'discount':
                ws.cell(row, 2).value = 'Discount'
            else:
                ws.cell(row, 2).value = 'Other Charges'
            
            ws.cell(row, 3).value = data.get('entry_number')
            for c in range(4, 10):
                ws.cell(row, c).value = ''
            
            if entry_type == 'opening_balance':
                if amount >= 0:
                    running_balance += amount
                    ws.cell(row, 10).value = amount
                    ws.cell(row, 10).number_format = '₹#,##0.00'
                    ws.cell(row, 11).value = ''
                else:
                    running_balance -= abs(amount)
                    ws.cell(row, 10).value = ''
                    ws.cell(row, 11).value = abs(amount)
                    ws.cell(row, 11).number_format = '₹#,##0.00'
            elif entry_type in ['freight', 'discount'] or amount < 0:
                running_balance -= abs(amount)
                ws.cell(row, 10).value = ''
                ws.cell(row, 11).value = abs(amount)
                ws.cell(row, 11).number_format = '₹#,##0.00'
            else:
                running_balance += abs(amount)
                ws.cell(row, 10).value = abs(amount)
                ws.cell(row, 10).number_format = '₹#,##0.00'
                ws.cell(row, 11).value = ''
            
            ws.cell(row, 12).value = running_balance
            ws.cell(row, 12).number_format = '₹#,##0.00'
            
            for c in range(1, 13):
                ws.cell(row, c).border = border
            row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"customer_ledger_{customer.get('name', 'unknown').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/reports/customer-ledger/{customer_id}/export-pdf")
async def export_customer_ledger_pdf(customer_id: str):
    """Export customer ledger to PDF - Each item with individual amount and balance"""
    ledger_response = await get_customer_ledger(customer_id)
    customer = ledger_response['customer']
    summary = ledger_response['summary']
    
    pdf_buffer = io.BytesIO()
    # Use landscape A4 for better width
    from reportlab.lib.pagesizes import landscape
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), leftMargin=15, rightMargin=15, topMargin=20, bottomMargin=20)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16,
                                 textColor=colors.HexColor('#1a1a1a'), spaceAfter=10, alignment=TA_CENTER)
    
    title = Paragraph(f"<b>Customer Ledger - {customer.get('name', 'N/A')}</b>", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.15*inch))
    
    customer_info = [
        ['Customer ID:', customer.get('id', 'N/A')],
        ['Address:', customer.get('address', 'N/A')],
        ['Phone:', customer.get('phone', 'N/A')],
        ['GST:', customer.get('gst_number', 'N/A')]
    ]
    
    customer_table = Table(customer_info, colWidths=[1*inch, 3*inch])
    customer_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 0.15*inch))
    
    summary_data = [
        ['Summary', ''],
        ['Total Invoiced:', f"₹{summary.get('total_invoiced', 0):,.0f}"],
        ['Total Paid:', f"₹{summary.get('total_paid', 0):,.0f}"],
        ['Outstanding:', f"₹{summary.get('net_balance', 0):,.0f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[1.5*inch, 1.2*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, -1), (1, -1), colors.red),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.2*inch))
    
    elements.append(Paragraph("<b>Transactions</b>", styles['Heading3']))
    elements.append(Spacer(1, 0.08*inch))
    
    # Transaction table
    txn_data = [['Date', 'Type', 'Ref', 'Product', 'Qty', 'Rate', 'Disc%', 'GST%', 'Tax', 'Debit', 'Credit', 'Balance']]
    
    running_balance = 0
    all_txns = []
    
    for inv in ledger_response['invoices']:
        all_txns.append(('invoice', inv.get('invoice_date'), inv))
    for cn in ledger_response['credit_notes']:
        all_txns.append(('credit_note', cn.get('credit_note_date'), cn))
    for pmt in ledger_response['payments']:
        all_txns.append(('payment', pmt.get('payment_date'), pmt))
    for je in ledger_response['journal_entries']:
        all_txns.append(('journal', je.get('entry_date'), je))
    
    all_txns.sort(key=lambda x: x[1] if x[1] else datetime.min)
    
    for txn_type, txn_date, data in all_txns:
        date_str = txn_date.strftime('%d/%m/%y') if isinstance(txn_date, datetime) else str(txn_date)
        
        if txn_type == 'invoice':
            if data.get('items'):
                for idx, item in enumerate(data.get('items', [])):
                    item_total = item.get('quantity', 0) * item.get('price', 0)
                    disc_amt = item_total * (item.get('discount_percent', 0) / 100)
                    taxable = item_total - disc_amt
                    tax_amt = taxable * (item.get('gst_rate', 0) / 100)
                    item_final = taxable + tax_amt
                    
                    # Update balance for each item
                    running_balance += item_final
                    
                    row = [
                        date_str if idx == 0 else '',
                        'Invoice',
                        data.get('invoice_number', ''),
                        item.get('product_name', ''),
                        f"{item.get('quantity', 0)}",
                        f"₹{item.get('price', 0):.0f}",
                        f"{item.get('discount_percent', 0)}%",
                        f"{item.get('gst_rate', 0)}%",
                        f"₹{tax_amt:.0f}",
                        f"₹{item_final:,.0f}",
                        '',
                        f"₹{running_balance:,.0f}"
                    ]
                    txn_data.append(row)
            else:
                invoice_total = data.get('grand_total', 0)
                running_balance += invoice_total
                txn_data.append([
                    date_str, 'Invoice', data.get('invoice_number', ''),
                    '', '', '', '', '', '',
                    f"₹{invoice_total:,.0f}", '', f"₹{running_balance:,.0f}"
                ])
        
        elif txn_type == 'credit_note':
            if data.get('items'):
                for idx, item in enumerate(data.get('items', [])):
                    item_total = item.get('quantity', 0) * item.get('price', 0)
                    disc_amt = item_total * (item.get('discount_percent', 0) / 100)
                    taxable = item_total - disc_amt
                    tax_amt = taxable * (item.get('gst_rate', 0) / 100)
                    item_final = taxable + tax_amt
                    
                    running_balance -= item_final
                    
                    row = [
                        date_str if idx == 0 else '',
                        'Credit Note',
                        data.get('credit_note_number', ''),
                        item.get('product_name', ''),
                        f"{item.get('quantity', 0)}",
                        f"₹{item.get('price', 0):.0f}",
                        f"{item.get('discount_percent', 0)}%",
                        f"{item.get('gst_rate', 0)}%",
                        f"₹{tax_amt:.0f}",
                        '',
                        f"₹{item_final:,.0f}",
                        f"₹{running_balance:,.0f}"
                    ]
                    txn_data.append(row)
            else:
                credit_amt = data.get('credit_amount', 0)
                running_balance -= credit_amt
                txn_data.append([
                    date_str, 'Credit Note', data.get('credit_note_number', ''),
                    '', '', '', '', '', '',
                    '', f"₹{credit_amt:,.0f}", f"₹{running_balance:,.0f}"
                ])
        
        elif txn_type == 'payment':
            pmt_amt = data.get('payment_amount', 0)
            running_balance -= pmt_amt
            txn_data.append([
                date_str, 'Payment', data.get('payment_number', ''),
                '', '', '', '', '', '',
                '', f"₹{pmt_amt:,.0f}", f"₹{running_balance:,.0f}"
            ])
        
        elif txn_type == 'journal':
            amount = data.get('amount', 0)
            entry_type = data.get('entry_type', '')
            
            if entry_type == 'opening_balance':
                txn_type_label = 'Opening Bal'
                if amount >= 0:
                    running_balance += amount
                    debit_val = f"₹{amount:,.0f}"
                    credit_val = ''
                else:
                    running_balance -= abs(amount)
                    debit_val = ''
                    credit_val = f"₹{abs(amount):,.0f}"
            elif entry_type == 'freight':
                txn_type_label = 'Freight'
                running_balance -= abs(amount)
                debit_val = ''
                credit_val = f"₹{abs(amount):,.0f}"
            elif entry_type == 'discount':
                txn_type_label = 'Discount'
                running_balance -= abs(amount)
                debit_val = ''
                credit_val = f"₹{abs(amount):,.0f}"
            else:
                txn_type_label = 'Other'
                running_balance += abs(amount)
                debit_val = f"₹{abs(amount):,.0f}"
                credit_val = ''
            
            txn_data.append([
                date_str, txn_type_label, data.get('entry_number', ''),
                '', '', '', '', '', '',
                debit_val, credit_val, f"₹{running_balance:,.0f}"
            ])
    
    col_widths = [0.6*inch, 0.6*inch, 0.75*inch, 2.2*inch, 0.45*inch, 0.6*inch, 0.45*inch, 0.45*inch, 0.55*inch, 0.7*inch, 0.7*inch, 0.75*inch]
    txn_table = Table(txn_data, colWidths=col_widths, repeatRows=1)
    txn_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7.5),
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (3, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(txn_table)
    
    doc.build(elements)
    pdf_buffer.seek(0)
    
    filename = f"customer_ledger_{customer.get('name', 'unknown').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ========== HR MODULE API ROUTES ==========

# Helper function to generate employee code
async def generate_employee_code():
    """Generate employee code in format EMP-XXXX"""
    count = await db.employees.count_documents({})
    return f"EMP-{str(count + 1).zfill(4)}"


# Department Routes
@api_router.post("/hr/departments", response_model=Department)
async def create_department(department: DepartmentCreate):
    dept_obj = Department(**department.model_dump())
    doc = dept_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.departments.insert_one(doc)
    return dept_obj

@api_router.get("/hr/departments", response_model=List[Department])
async def get_departments():
    departments = await db.departments.find({}, {"_id": 0}).to_list(1000)
    for dept in departments:
        if isinstance(dept.get('created_at'), str):
            dept['created_at'] = datetime.fromisoformat(dept['created_at'])
    return departments

@api_router.get("/hr/departments/{department_id}", response_model=Department)
async def get_department(department_id: str):
    department = await db.departments.find_one({"id": department_id}, {"_id": 0})
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")
    if isinstance(department.get('created_at'), str):
        department['created_at'] = datetime.fromisoformat(department['created_at'])
    return department

@api_router.put("/hr/departments/{department_id}")
async def update_department(department_id: str, department_update: dict):
    result = await db.departments.update_one(
        {"id": department_id},
        {"$set": department_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    return {"message": "Department updated successfully"}

@api_router.delete("/hr/departments/{department_id}")
async def delete_department(department_id: str):
    result = await db.departments.delete_one({"id": department_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    return {"message": "Department deleted successfully"}


# Employee Routes
@api_router.post("/hr/employees", response_model=Employee)
async def create_employee(employee: EmployeeCreate):
    employee_code = await generate_employee_code()
    employee_obj = Employee(employee_code=employee_code, **employee.model_dump())
    doc = employee_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('date_of_joining'):
        doc['date_of_joining'] = doc['date_of_joining'].isoformat()
    if doc.get('date_of_birth'):
        doc['date_of_birth'] = doc['date_of_birth'].isoformat()
    await db.employees.insert_one(doc)
    return employee_obj

@api_router.get("/hr/employees", response_model=List[Employee])
async def get_employees():
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    for emp in employees:
        if isinstance(emp.get('created_at'), str):
            emp['created_at'] = datetime.fromisoformat(emp['created_at'])
        if isinstance(emp.get('date_of_joining'), str):
            emp['date_of_joining'] = datetime.fromisoformat(emp['date_of_joining'])
        if emp.get('date_of_birth') and isinstance(emp.get('date_of_birth'), str):
            emp['date_of_birth'] = datetime.fromisoformat(emp['date_of_birth'])
    return employees

@api_router.get("/hr/employees/{employee_id}", response_model=Employee)
async def get_employee(employee_id: str):
    employee = await db.employees.find_one({"id": employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    if isinstance(employee.get('created_at'), str):
        employee['created_at'] = datetime.fromisoformat(employee['created_at'])
    if isinstance(employee.get('date_of_joining'), str):
        employee['date_of_joining'] = datetime.fromisoformat(employee['date_of_joining'])
    if employee.get('date_of_birth') and isinstance(employee.get('date_of_birth'), str):
        employee['date_of_birth'] = datetime.fromisoformat(employee['date_of_birth'])
    return employee

@api_router.put("/hr/employees/{employee_id}")
async def update_employee(employee_id: str, employee_update: dict):
    result = await db.employees.update_one(
        {"id": employee_id},
        {"$set": employee_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee updated successfully"}

@api_router.delete("/hr/employees/{employee_id}")
async def delete_employee(employee_id: str):
    result = await db.employees.delete_one({"id": employee_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee deleted successfully"}


# Attendance Routes
@api_router.post("/hr/attendance", response_model=Attendance)
async def mark_attendance(attendance: AttendanceCreate):
    attendance_obj = Attendance(**attendance.model_dump())
    doc = attendance_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['date'] = doc['date'].isoformat()
    if doc.get('check_in'):
        doc['check_in'] = doc['check_in'].isoformat()
    if doc.get('check_out'):
        doc['check_out'] = doc['check_out'].isoformat()
    await db.attendance.insert_one(doc)
    return attendance_obj

@api_router.get("/hr/attendance")
async def get_attendance(employee_id: Optional[str] = None, date: Optional[str] = None):
    query = {}
    if employee_id:
        query['employee_id'] = employee_id
    if date:
        query['date'] = {"$regex": f"^{date}"}
    
    attendance_records = await db.attendance.find(query, {"_id": 0}).to_list(1000)
    return attendance_records

@api_router.put("/hr/attendance/{attendance_id}")
async def update_attendance(attendance_id: str, attendance_update: dict):
    result = await db.attendance.update_one(
        {"id": attendance_id},
        {"$set": attendance_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    return {"message": "Attendance updated successfully"}


# Leave Type Routes
@api_router.post("/hr/leave-types", response_model=LeaveType)
async def create_leave_type(leave_type: LeaveTypeCreate):
    leave_type_obj = LeaveType(**leave_type.model_dump())
    doc = leave_type_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.leave_types.insert_one(doc)
    return leave_type_obj

@api_router.get("/hr/leave-types", response_model=List[LeaveType])
async def get_leave_types():
    leave_types = await db.leave_types.find({}, {"_id": 0}).to_list(100)
    for lt in leave_types:
        if isinstance(lt.get('created_at'), str):
            lt['created_at'] = datetime.fromisoformat(lt['created_at'])
    return leave_types


# Leave Request Routes
@api_router.post("/hr/leave-requests", response_model=LeaveRequest)
async def create_leave_request(leave_request: LeaveRequestCreate):
    leave_request_obj = LeaveRequest(**leave_request.model_dump())
    doc = leave_request_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['start_date'] = doc['start_date'].isoformat()
    doc['end_date'] = doc['end_date'].isoformat()
    await db.leave_requests.insert_one(doc)
    return leave_request_obj

@api_router.get("/hr/leave-requests")
async def get_leave_requests(employee_id: Optional[str] = None, status: Optional[str] = None):
    query = {}
    if employee_id:
        query['employee_id'] = employee_id
    if status:
        query['status'] = status
    
    leave_requests = await db.leave_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return leave_requests

@api_router.put("/hr/leave-requests/{request_id}/approve")
async def approve_leave_request(request_id: str, approval_data: dict):
    approved_by = approval_data.get("approved_by", "Manager")
    result = await db.leave_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": approved_by,
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Leave request not found")
    return {"message": "Leave request approved"}

@api_router.put("/hr/leave-requests/{request_id}/reject")
async def reject_leave_request(request_id: str, rejection_data: dict):
    rejection_reason = rejection_data.get("rejection_reason", "")
    result = await db.leave_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "rejected",
            "rejection_reason": rejection_reason
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Leave request not found")
    return {"message": "Leave request rejected"}


# Payslip Routes
@api_router.post("/hr/payslips", response_model=Payslip)
async def create_payslip(payslip: PayslipCreate):
    # Calculate totals
    total_earnings = payslip.basic_salary
    total_deductions = 0.0
    
    for component in payslip.components:
        if component.component_type == "earning":
            total_earnings += component.amount
        else:
            total_deductions += component.amount
    
    net_salary = total_earnings - total_deductions
    
    payslip_obj = Payslip(
        **payslip.model_dump(),
        total_earnings=total_earnings,
        total_deductions=total_deductions,
        net_salary=net_salary
    )
    
    doc = payslip_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.payslips.insert_one(doc)
    return payslip_obj

@api_router.get("/hr/payslips")
async def get_payslips(employee_id: Optional[str] = None, month: Optional[int] = None, year: Optional[int] = None):
    query = {}
    if employee_id:
        query['employee_id'] = employee_id
    if month:
        query['month'] = month
    if year:
        query['year'] = year
    
    payslips = await db.payslips.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return payslips

@api_router.put("/hr/payslips/{payslip_id}")
async def update_payslip(payslip_id: str, payslip_update: dict):
    result = await db.payslips.update_one(
        {"id": payslip_id},
        {"$set": payslip_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Payslip not found")
    return {"message": "Payslip updated successfully"}


# Performance Review Routes
@api_router.post("/hr/performance-reviews", response_model=PerformanceReview)
async def create_performance_review(review: PerformanceReviewCreate):
    review_obj = PerformanceReview(**review.model_dump())
    doc = review_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['review_period_start'] = doc['review_period_start'].isoformat()
    doc['review_period_end'] = doc['review_period_end'].isoformat()
    await db.performance_reviews.insert_one(doc)
    return review_obj

@api_router.get("/hr/performance-reviews")
async def get_performance_reviews(employee_id: Optional[str] = None):
    query = {}
    if employee_id:
        query['employee_id'] = employee_id
    
    reviews = await db.performance_reviews.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return reviews

@api_router.put("/hr/performance-reviews/{review_id}")
async def update_performance_review(review_id: str, review_update: dict):
    result = await db.performance_reviews.update_one(
        {"id": review_id},
        {"$set": review_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Performance review not found")
    return {"message": "Performance review updated successfully"}


# Goal Routes
@api_router.post("/hr/goals", response_model=Goal)
async def create_goal(goal: GoalCreate):
    goal_obj = Goal(**goal.model_dump())
    doc = goal_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['target_date'] = doc['target_date'].isoformat()
    await db.goals.insert_one(doc)
    return goal_obj

@api_router.get("/hr/goals")
async def get_goals(employee_id: Optional[str] = None):
    query = {}
    if employee_id:
        query['employee_id'] = employee_id
    
    goals = await db.goals.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return goals

@api_router.put("/hr/goals/{goal_id}")
async def update_goal(goal_id: str, goal_update: dict):
    result = await db.goals.update_one(
        {"id": goal_id},
        {"$set": goal_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Goal not found")
    return {"message": "Goal updated successfully"}


# Employee Document Routes
@api_router.post("/hr/documents", response_model=EmployeeDocument)
async def create_employee_document(document: EmployeeDocumentCreate):
    document_obj = EmployeeDocument(**document.model_dump())
    doc = document_obj.model_dump()
    doc['uploaded_at'] = doc['uploaded_at'].isoformat()
    await db.employee_documents.insert_one(doc)
    return document_obj

@api_router.get("/hr/documents")
async def get_employee_documents(employee_id: Optional[str] = None):
    query = {}
    if employee_id:
        query['employee_id'] = employee_id
    
    documents = await db.employee_documents.find(query, {"_id": 0}).sort("uploaded_at", -1).to_list(1000)
    return documents

@api_router.delete("/hr/documents/{document_id}")
async def delete_employee_document(document_id: str):
    result = await db.employee_documents.delete_one({"id": document_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": "Document deleted successfully"}


# Job Posting Routes
@api_router.post("/hr/job-postings", response_model=JobPosting)
async def create_job_posting(job: JobPostingCreate):
    job_obj = JobPosting(**job.model_dump())
    doc = job_obj.model_dump()
    doc['posted_date'] = doc['posted_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    if doc.get('closing_date'):
        doc['closing_date'] = doc['closing_date'].isoformat()
    await db.job_postings.insert_one(doc)
    return job_obj

@api_router.get("/hr/job-postings", response_model=List[JobPosting])
async def get_job_postings(status: Optional[str] = None):
    query = {}
    if status:
        query['status'] = status
    
    jobs = await db.job_postings.find(query, {"_id": 0}).sort("posted_date", -1).to_list(1000)
    for job in jobs:
        if isinstance(job.get('posted_date'), str):
            job['posted_date'] = datetime.fromisoformat(job['posted_date'])
        if isinstance(job.get('created_at'), str):
            job['created_at'] = datetime.fromisoformat(job['created_at'])
        if job.get('closing_date') and isinstance(job.get('closing_date'), str):
            job['closing_date'] = datetime.fromisoformat(job['closing_date'])
    return jobs

@api_router.put("/hr/job-postings/{job_id}")
async def update_job_posting(job_id: str, job_update: dict):
    result = await db.job_postings.update_one(
        {"id": job_id},
        {"$set": job_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Job posting not found")
    return {"message": "Job posting updated successfully"}


# Candidate Routes
@api_router.post("/hr/candidates", response_model=Candidate)
async def create_candidate(candidate: CandidateCreate):
    candidate_obj = Candidate(**candidate.model_dump())
    doc = candidate_obj.model_dump()
    doc['applied_date'] = doc['applied_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.candidates.insert_one(doc)
    return candidate_obj

@api_router.get("/hr/candidates")
async def get_candidates(job_posting_id: Optional[str] = None, status: Optional[str] = None):
    query = {}
    if job_posting_id:
        query['job_posting_id'] = job_posting_id
    if status:
        query['status'] = status
    
    candidates = await db.candidates.find(query, {"_id": 0}).sort("applied_date", -1).to_list(1000)
    return candidates

@api_router.put("/hr/candidates/{candidate_id}")
async def update_candidate(candidate_id: str, candidate_update: dict):
    result = await db.candidates.update_one(
        {"id": candidate_id},
        {"$set": candidate_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Candidate not found")
    return {"message": "Candidate updated successfully"}


# Interview Routes
@api_router.post("/hr/interviews", response_model=Interview)
async def create_interview(interview: InterviewCreate):
    interview_obj = Interview(**interview.model_dump())
    doc = interview_obj.model_dump()
    doc['interview_date'] = doc['interview_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.interviews.insert_one(doc)
    return interview_obj

@api_router.get("/hr/interviews")
async def get_interviews(candidate_id: Optional[str] = None, interviewer_id: Optional[str] = None):
    query = {}
    if candidate_id:
        query['candidate_id'] = candidate_id
    if interviewer_id:
        query['interviewer_id'] = interviewer_id
    
    interviews = await db.interviews.find(query, {"_id": 0}).sort("interview_date", 1).to_list(1000)
    return interviews

@api_router.put("/hr/interviews/{interview_id}")
async def update_interview(interview_id: str, interview_update: dict):
    result = await db.interviews.update_one(
        {"id": interview_id},
        {"$set": interview_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Interview not found")
    return {"message": "Interview updated successfully"}


# HR Dashboard Stats
@api_router.get("/hr/dashboard/stats")
async def get_hr_dashboard_stats():
    total_employees = await db.employees.count_documents({"status": "active"})
    total_departments = await db.departments.count_documents({})
    pending_leaves = await db.leave_requests.count_documents({"status": "pending"})
    active_jobs = await db.job_postings.count_documents({"status": "active"})
    total_candidates = await db.candidates.count_documents({})
    
    # Get attendance today
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    present_today = await db.attendance.count_documents({
        "date": {"$regex": f"^{today}"},
        "status": "present"
    })
    
    return {
        "total_employees": total_employees,
        "total_departments": total_departments,
        "pending_leaves": pending_leaves,
        "active_job_postings": active_jobs,
        "total_candidates": total_candidates,
        "present_today": present_today
    }


# Inventory Dashboard Stats
@api_router.get("/inventory/dashboard/stats")
async def get_inventory_dashboard_stats():
    total_raw_materials = await db.raw_materials.count_documents({})
    total_finished_goods = await db.finished_goods.count_documents({})
    total_packing_materials = await db.packing_materials.count_documents({})
    low_stock_items = await db.raw_materials.count_documents({"quantity": {"$lt": 100}})
    pending_approvals = await db.purchase_requests.count_documents({"approval_status": "pending"})
    total_stock_value = 0
    
    # Calculate total stock value
    raw_materials = await db.raw_materials.find({}).to_list(1000)
    for item in raw_materials:
        total_stock_value += item.get('quantity', 0) * item.get('rate', 0)
    
    return {
        "total_raw_materials": total_raw_materials,
        "total_finished_goods": total_finished_goods,
        "total_packing_materials": total_packing_materials,
        "low_stock_items": low_stock_items,
        "pending_approvals": pending_approvals,
        "total_stock_value": round(total_stock_value, 2)
    }


# Manufacturing Dashboard Stats
@api_router.get("/manufacturing/dashboard/stats")
async def get_manufacturing_dashboard_stats():
    total_boms = await db.boms.count_documents({})
    active_production_orders = await db.production_orders.count_documents({"status": "in_progress"})
    pending_production_orders = await db.production_orders.count_documents({"status": "pending"})
    completed_today = await db.production_orders.count_documents({
        "status": "completed",
        "completion_date": {"$gte": datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()}
    })
    
    return {
        "total_boms": total_boms,
        "active_production_orders": active_production_orders,
        "pending_production_orders": pending_production_orders,
        "completed_today": completed_today,
        "total_production_orders": await db.production_orders.count_documents({})
    }


# Sales Dashboard Stats
@api_router.get("/sales/dashboard/stats")
async def get_sales_dashboard_stats():
    total_customers = await db.customers.count_documents({})
    total_quotations = await db.quotations.count_documents({})
    pending_quotations = await db.quotations.count_documents({"status": "pending"})
    total_sales_orders = await db.sales_orders.count_documents({})
    pending_invoices = await db.invoices.count_documents({"status": {"$in": ["pending", "sent"]}})
    
    # Calculate total sales value
    invoices = await db.invoices.find({"status": "paid"}).to_list(1000)
    total_revenue = sum([inv.get('total_amount', 0) for inv in invoices])
    
    return {
        "total_customers": total_customers,
        "total_quotations": total_quotations,
        "pending_quotations": pending_quotations,
        "total_sales_orders": total_sales_orders,
        "pending_invoices": pending_invoices,
        "total_revenue": round(total_revenue, 2)
    }


# Purchase Dashboard Stats
@api_router.get("/purchase/dashboard/stats")
async def get_purchase_dashboard_stats():
    total_suppliers = await db.suppliers.count_documents({})
    pending_purchase_requests = await db.purchase_requests.count_documents({"approval_status": "pending"})
    total_purchase_orders = await db.purchase_orders.count_documents({})
    pending_purchase_orders = await db.purchase_orders.count_documents({"status": {"$in": ["pending", "approved"]}})
    pending_grns = await db.grns.count_documents({"status": "pending"})
    
    # Calculate total purchase value
    purchase_orders = await db.purchase_orders.find({"status": "completed"}).to_list(1000)
    total_purchase_value = sum([po.get('total_amount', 0) for po in purchase_orders])
    
    return {
        "total_suppliers": total_suppliers,
        "pending_purchase_requests": pending_purchase_requests,
        "total_purchase_orders": total_purchase_orders,
        "pending_purchase_orders": pending_purchase_orders,
        "pending_grns": pending_grns,
        "total_purchase_value": round(total_purchase_value, 2)
    }


# Finance Dashboard Stats
@api_router.get("/finance/dashboard/stats")
async def get_finance_dashboard_stats():
    # Accounts receivable (pending customer invoices)
    customer_invoices = await db.invoices.find({"status": {"$in": ["pending", "sent"]}}).to_list(1000)
    accounts_receivable = sum([inv.get('total_amount', 0) for inv in customer_invoices])
    
    # Accounts payable (pending supplier bills)
    supplier_bills = await db.supplier_bills.find({"status": "pending"}).to_list(1000)
    accounts_payable = sum([bill.get('total_amount', 0) for bill in supplier_bills])
    
    # Total revenue (paid invoices)
    paid_invoices = await db.invoices.find({"status": "paid"}).to_list(1000)
    total_revenue = sum([inv.get('total_amount', 0) for inv in paid_invoices])
    
    # Total expenses (paid bills)
    paid_bills = await db.supplier_bills.find({"status": "paid"}).to_list(1000)
    total_expenses = sum([bill.get('total_amount', 0) for bill in paid_bills])
    
    profit = total_revenue - total_expenses
    
    return {
        "accounts_receivable": round(accounts_receivable, 2),
        "accounts_payable": round(accounts_payable, 2),
        "total_revenue": round(total_revenue, 2),
        "total_expenses": round(total_expenses, 2),
        "profit": round(profit, 2),
        "pending_invoices": len(customer_invoices),
        "pending_bills": len(supplier_bills)
    }


# ========== USER & ROLE MANAGEMENT ROUTES ==========

# Role Management
@api_router.post("/admin/roles", response_model=Role)
async def create_role(role: RoleCreate):
    role_obj = Role(**role.model_dump())
    doc = role_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.roles.insert_one(doc)
    return role_obj

@api_router.get("/admin/roles", response_model=List[Role])
async def get_roles():
    roles = await db.roles.find({}, {"_id": 0}).to_list(1000)
    for role in roles:
        if isinstance(role.get('created_at'), str):
            role['created_at'] = datetime.fromisoformat(role['created_at'])
    return roles

@api_router.get("/admin/roles/{role_id}", response_model=Role)
async def get_role(role_id: str):
    role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    if isinstance(role.get('created_at'), str):
        role['created_at'] = datetime.fromisoformat(role['created_at'])
    return role

@api_router.put("/admin/roles/{role_id}")
async def update_role(role_id: str, role_update: dict):
    result = await db.roles.update_one(
        {"id": role_id},
        {"$set": role_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Role not found")
    return {"message": "Role updated successfully"}

@api_router.delete("/admin/roles/{role_id}")
async def delete_role(role_id: str):
    result = await db.roles.delete_one({"id": role_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Role not found")
    return {"message": "Role deleted successfully"}


# User Management
@api_router.post("/admin/users", response_model=UserWithRole)
async def create_user(user: UserCreate):
    # Check if username already exists
    existing_user = await db.users.find_one({"username": user.username})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Create user with hashed password
    hashed_password = pwd_context.hash(user.password)
    user_obj = UserWithRole(**user.model_dump(exclude={'password'}))
    doc = user_obj.model_dump()
    doc['hashed_password'] = hashed_password
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    return user_obj

@api_router.get("/admin/users", response_model=List[UserWithRole])
async def get_users():
    users = await db.users.find({}, {"_id": 0, "hashed_password": 0}).to_list(1000)
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    return users

@api_router.get("/admin/users/{user_id}", response_model=UserWithRole)
async def get_user(user_id: str):
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "hashed_password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if isinstance(user.get('created_at'), str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    return user

@api_router.put("/admin/users/{user_id}")
async def update_user(user_id: str, user_update: dict):
    # If password is being updated, hash it
    if 'password' in user_update:
        user_update['hashed_password'] = pwd_context.hash(user_update['password'])
        del user_update['password']
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": user_update}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User updated successfully"}

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str):
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}

@api_router.post("/admin/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, password_data: dict):
    new_password = password_data.get('password')
    if not new_password:
        raise HTTPException(status_code=400, detail="Password is required")
    
    hashed_password = pwd_context.hash(new_password)
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"hashed_password": hashed_password}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "Password reset successfully"}


app.include_router(api_router)
app.include_router(whatsapp_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()# Invoice Management Backend APIs
# Add this section to your server.py

from typing import List, Optional
from datetime import datetime, timezone
from uuid import uuid4
import os
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Invoice Models
class InvoiceItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    item_name: str
    description: Optional[str] = ""
    quantity: float
    unit_price: float
    tax_rate: Optional[float] = 0
    amount: float

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    invoice_number: str
    invoice_date: str
    due_date: str
    
    # Customer/Dealer Info
    customer_id: Optional[str] = None
    customer_name: str
    customer_phone: str
    customer_email: Optional[str] = ""
    customer_address: Optional[str] = ""
    customer_gstin: Optional[str] = ""
    
    # Invoice Items
    items: List[InvoiceItem]
    
    # Amounts
    subtotal: float
    tax_amount: float
    discount_amount: Optional[float] = 0
    total_amount: float
    
    # Payment Info
    payment_terms: Optional[str] = "Net 30"
    payment_status: str = "unpaid"  # unpaid, partially_paid, paid
    paid_amount: Optional[float] = 0
    
    # WhatsApp
    send_whatsapp: Optional[bool] = False
    whatsapp_sent: Optional[bool] = False
    whatsapp_sent_at: Optional[str] = None
    
    # Metadata
    notes: Optional[str] = ""
    status: str = "draft"  # draft, sent, paid, cancelled
    created_by: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: Optional[str] = None

class InvoiceCreate(BaseModel):
    invoice_number: str
    invoice_date: str
    due_date: str
    customer_name: str
    customer_phone: str
    customer_email: Optional[str] = ""
    customer_address: Optional[str] = ""
    customer_gstin: Optional[str] = ""
    items: List[InvoiceItem]
    subtotal: float
    tax_amount: float
    discount_amount: Optional[float] = 0
    total_amount: float
    payment_terms: Optional[str] = "Net 30"
    notes: Optional[str] = ""
    send_whatsapp: Optional[bool] = False

# PDF Generation Function
async def generate_invoice_pdf(invoice: Invoice) -> str:
    """Generate PDF for invoice and return file path"""
    try:
        # Create invoices directory if doesn't exist
        invoices_dir = Path("/app/invoices")
        invoices_dir.mkdir(exist_ok=True)
        
        pdf_filename = f"Invoice_{invoice.invoice_number}.pdf"
        pdf_path = invoices_dir / pdf_filename
        
        # Create PDF
        doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a56db'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("INVOICE", title_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Company Info & Invoice Details (side by side)
        company_invoice_data = [
            [
                Paragraph("<b>Nectar</b><br/>Your Business Address<br/>Phone: +91-XXXXXXXXXX<br/>Email: info@nectar.com", styles['Normal']),
                Paragraph(f"<b>Invoice #:</b> {invoice.invoice_number}<br/><b>Date:</b> {invoice.invoice_date[:10]}<br/><b>Due Date:</b> {invoice.due_date[:10]}", styles['Normal'])
            ]
        ]
        
        company_invoice_table = Table(company_invoice_data, colWidths=[3*inch, 3*inch])
        company_invoice_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        story.append(company_invoice_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Bill To
        story.append(Paragraph("<b>Bill To:</b>", styles['Heading3']))
        bill_to_text = f"{invoice.customer_name}<br/>"
        if invoice.customer_address:
            bill_to_text += f"{invoice.customer_address}<br/>"
        bill_to_text += f"Phone: {invoice.customer_phone}<br/>"
        if invoice.customer_email:
            bill_to_text += f"Email: {invoice.customer_email}<br/>"
        if invoice.customer_gstin:
            bill_to_text += f"GSTIN: {invoice.customer_gstin}"
        
        story.append(Paragraph(bill_to_text, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Items Table
        items_data = [['Item', 'Description', 'Qty', 'Rate', 'Tax', 'Amount']]
        
        for item in invoice.items:
            items_data.append([
                item.item_name,
                item.description or '-',
                str(item.quantity),
                f"₹{item.unit_price:,.2f}",
                f"{item.tax_rate}%",
                f"₹{item.amount:,.2f}"
            ])
        
        items_table = Table(items_data, colWidths=[1.5*inch, 2*inch, 0.7*inch, 1*inch, 0.7*inch, 1*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ]))
        story.append(items_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Totals
        totals_data = [
            ['Subtotal:', f"₹{invoice.subtotal:,.2f}"],
            ['Tax:', f"₹{invoice.tax_amount:,.2f}"],
        ]
        
        if invoice.discount_amount and invoice.discount_amount > 0:
            totals_data.append(['Discount:', f"-₹{invoice.discount_amount:,.2f}"])
        
        totals_data.append(['<b>Total Amount:</b>', f"<b>₹{invoice.total_amount:,.2f}</b>"])
        
        totals_table = Table(totals_data, colWidths=[4.5*inch, 1.5*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1a56db')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1a56db')),
        ]))
        story.append(totals_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Payment Terms & Notes
        if invoice.payment_terms:
            story.append(Paragraph(f"<b>Payment Terms:</b> {invoice.payment_terms}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
        
        if invoice.notes:
            story.append(Paragraph(f"<b>Notes:</b><br/>{invoice.notes}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
        
        # Footer
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph("Thank you for your business!", footer_style))
        
        # Build PDF
        doc.build(story)
        
        return str(pdf_path)
        
    except Exception as e:
        print(f"Error generating PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {str(e)}")

# Invoice Routes
@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(invoice_data: InvoiceCreate):
    """Create a new invoice"""
    try:
        invoice = Invoice(**invoice_data.dict())
        invoice.status = "sent"
        
        # Generate PDF
        pdf_path = await generate_invoice_pdf(invoice)
        
        # Save to database
        await db.invoices.insert_one(invoice.dict())
        
        # Send via WhatsApp if requested
        if invoice.send_whatsapp and invoice.customer_phone:
            try:
                async with httpx.AsyncClient() as client:
                    whatsapp_data = {
                        "phone_number": invoice.customer_phone,
                        "invoice_number": invoice.invoice_number,
                        "dealer_name": invoice.customer_name,
                        "amount": str(invoice.total_amount),
                        "due_date": invoice.due_date[:10],
                        "pdf_path": pdf_path
                    }
                    
                    response = await client.post(
                        "http://localhost:3001/send-invoice",
                        json=whatsapp_data,
                        timeout=60.0
                    )
                    
                    if response.json().get('success'):
                        invoice.whatsapp_sent = True
                        invoice.whatsapp_sent_at = datetime.now(timezone.utc).isoformat()
                        
                        # Update in database
                        await db.invoices.update_one(
                            {"id": invoice.id},
                            {"$set": {
                                "whatsapp_sent": True,
                                "whatsapp_sent_at": invoice.whatsapp_sent_at
                            }}
                        )
            except Exception as e:
                print(f"WhatsApp send failed: {e}")
                # Continue even if WhatsApp fails
        
        return invoice
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create invoice: {str(e)}")

@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices(
    status: Optional[str] = None,
    customer_name: Optional[str] = None,
    limit: int = 100
):
    """Get all invoices with optional filters"""
    try:
        query = {}
        
        if status:
            query["status"] = status
        
        if customer_name:
            query["customer_name"] = {"$regex": customer_name, "$options": "i"}
        
        invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
        return invoices
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/invoices/{invoice_id}", response_model=Invoice)
async def get_invoice(invoice_id: str):
    """Get invoice by ID"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return invoice

@api_router.put("/invoices/{invoice_id}", response_model=Invoice)
async def update_invoice(invoice_id: str, invoice_data: InvoiceCreate):
    """Update invoice"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    updated_data = invoice_data.dict()
    updated_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": updated_data}
    )
    
    updated_invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    return updated_invoice

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str):
    """Delete invoice"""
    result = await db.invoices.delete_one({"id": invoice_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"message": "Invoice deleted successfully"}

@api_router.post("/invoices/{invoice_id}/send-whatsapp")
async def resend_invoice_whatsapp(invoice_id: str):
    """Resend invoice via WhatsApp"""
    invoice_data = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice_data:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    invoice = Invoice(**invoice_data)
    
    # Find PDF
    pdf_path = f"/app/invoices/Invoice_{invoice.invoice_number}.pdf"
    
    if not os.path.exists(pdf_path):
        # Regenerate PDF if not found
        pdf_path = await generate_invoice_pdf(invoice)
    
    try:
        async with httpx.AsyncClient() as client:
            whatsapp_data = {
                "phone_number": invoice.customer_phone,
                "invoice_number": invoice.invoice_number,
                "dealer_name": invoice.customer_name,
                "amount": str(invoice.total_amount),
                "due_date": invoice.due_date[:10],
                "pdf_path": pdf_path
            }
            
            response = await client.post(
                "http://localhost:3001/send-invoice",
                json=whatsapp_data,
                timeout=60.0
            )
            
            if response.json().get('success'):
                # Update database
                await db.invoices.update_one(
                    {"id": invoice_id},
                    {"$set": {
                        "whatsapp_sent": True,
                        "whatsapp_sent_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                return {"success": True, "message": "Invoice sent via WhatsApp"}
            else:
                return {"success": False, "message": "Failed to send via WhatsApp"}
                
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send WhatsApp: {str(e)}")

@api_router.get("/invoices/{invoice_id}/pdf")
async def download_invoice_pdf(invoice_id: str):
    """Download invoice PDF"""
    invoice_data = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice_data:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    invoice = Invoice(**invoice_data)
    pdf_path = f"/app/invoices/Invoice_{invoice.invoice_number}.pdf"
    
    if not os.path.exists(pdf_path):
        # Generate PDF if not found
        pdf_path = await generate_invoice_pdf(invoice)
    
    from fastapi.responses import FileResponse
    return FileResponse(
        pdf_path,
        media_type='application/pdf',
        filename=f"Invoice_{invoice.invoice_number}.pdf"
    )
# Recovery Module Backend APIs
# Payment Recovery and Follow-up Management

from typing import List, Optional
from datetime import datetime, timezone, timedelta
from uuid import uuid4

# Models for Recovery
class PaymentRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    invoice_id: str
    invoice_number: str
    payment_date: str
    amount: float
    payment_method: str  # cash, cheque, bank_transfer, upi, etc.
    reference_number: Optional[str] = ""
    notes: Optional[str] = ""
    recorded_by: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class FollowUpNote(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid4()))
    invoice_id: str
    invoice_number: str
    follow_up_date: str
    contacted_person: Optional[str] = ""
    contact_method: str  # phone, email, whatsapp, visit
    notes: str
    next_follow_up_date: Optional[str] = None
    status: str  # contacted, promised_payment, disputed, no_response
    recorded_by: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Recovery Dashboard Stats
@api_router.get("/recovery/stats")
async def get_recovery_stats():
    """Get recovery dashboard statistics"""
    try:
        # Get all invoices
        all_invoices = await db.invoices.find({"status": {"$ne": "cancelled"}}, {"_id": 0}).to_list(1000)
        
        today = datetime.now(timezone.utc)
        
        stats = {
            "total_outstanding": 0,
            "overdue_count": 0,
            "overdue_amount": 0,
            "due_this_week": 0,
            "due_this_week_amount": 0,
            "partially_paid_count": 0,
            "partially_paid_amount": 0,
            "critical_overdue": 0,  # >30 days
            "critical_overdue_amount": 0,
            "by_age": {
                "0-7": {"count": 0, "amount": 0},
                "8-15": {"count": 0, "amount": 0},
                "16-30": {"count": 0, "amount": 0},
                "31-60": {"count": 0, "amount": 0},
                "60+": {"count": 0, "amount": 0}
            }
        }
        
        for invoice in all_invoices:
            if invoice.get('payment_status') == 'paid':
                continue
                
            due_date = datetime.fromisoformat(invoice['due_date'].replace('Z', '+00:00'))
            days_overdue = (today - due_date).days
            
            outstanding = invoice['total_amount'] - invoice.get('paid_amount', 0)
            stats['total_outstanding'] += outstanding
            
            # Overdue invoices
            if days_overdue > 0:
                stats['overdue_count'] += 1
                stats['overdue_amount'] += outstanding
                
                # Critical overdue (>30 days)
                if days_overdue > 30:
                    stats['critical_overdue'] += 1
                    stats['critical_overdue_amount'] += outstanding
                
                # Age buckets
                if days_overdue <= 7:
                    stats['by_age']['0-7']['count'] += 1
                    stats['by_age']['0-7']['amount'] += outstanding
                elif days_overdue <= 15:
                    stats['by_age']['8-15']['count'] += 1
                    stats['by_age']['8-15']['amount'] += outstanding
                elif days_overdue <= 30:
                    stats['by_age']['16-30']['count'] += 1
                    stats['by_age']['16-30']['amount'] += outstanding
                elif days_overdue <= 60:
                    stats['by_age']['31-60']['count'] += 1
                    stats['by_age']['31-60']['amount'] += outstanding
                else:
                    stats['by_age']['60+']['count'] += 1
                    stats['by_age']['60+']['amount'] += outstanding
            
            # Due this week
            if 0 <= days_overdue <= 7:
                stats['due_this_week'] += 1
                stats['due_this_week_amount'] += outstanding
            
            # Partially paid
            if invoice.get('payment_status') == 'partially_paid':
                stats['partially_paid_count'] += 1
                stats['partially_paid_amount'] += outstanding
        
        return stats
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Get Overdue Invoices
@api_router.get("/recovery/overdue-invoices")
async def get_overdue_invoices(
    days_overdue: Optional[int] = None,
    customer_name: Optional[str] = None,
    min_amount: Optional[float] = None
):
    """Get all overdue invoices with filters"""
    try:
        query = {
            "status": {"$ne": "cancelled"},
            "payment_status": {"$in": ["unpaid", "partially_paid"]}
        }
        
        if customer_name:
            query["customer_name"] = {"$regex": customer_name, "$options": "i"}
        
        invoices = await db.invoices.find(query, {"_id": 0}).to_list(1000)
        
        today = datetime.now(timezone.utc)
        overdue_invoices = []
        
        for invoice in invoices:
            due_date = datetime.fromisoformat(invoice['due_date'].replace('Z', '+00:00'))
            days_past = (today - due_date).days
            
            if days_past > 0:  # Only overdue
                outstanding = invoice['total_amount'] - invoice.get('paid_amount', 0)
                
                # Apply filters
                if days_overdue and days_past < days_overdue:
                    continue
                if min_amount and outstanding < min_amount:
                    continue
                
                invoice['days_overdue'] = days_past
                invoice['outstanding_amount'] = outstanding
                
                # Get last follow-up
                last_followup = await db.follow_ups.find_one(
                    {"invoice_id": invoice['id']},
                    {"_id": 0},
                    sort=[("created_at", -1)]
                )
                invoice['last_follow_up'] = last_followup
                
                # Get payment history
                payments = await db.payment_records.find(
                    {"invoice_id": invoice['id']},
                    {"_id": 0}
                ).to_list(100)
                invoice['payment_history'] = payments
                
                overdue_invoices.append(invoice)
        
        # Sort by days overdue (most critical first)
        overdue_invoices.sort(key=lambda x: x['days_overdue'], reverse=True)
        
        return overdue_invoices
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Record Payment
@api_router.post("/recovery/record-payment")
async def record_payment(payment: PaymentRecord):
    """Record a payment against an invoice"""
    try:
        # Get invoice
        invoice = await db.invoices.find_one({"id": payment.invoice_id}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        # Save payment record
        await db.payment_records.insert_one(payment.dict())
        
        # Update invoice
        paid_amount = invoice.get('paid_amount', 0) + payment.amount
        
        if paid_amount >= invoice['total_amount']:
            payment_status = 'paid'
        elif paid_amount > 0:
            payment_status = 'partially_paid'
        else:
            payment_status = 'unpaid'
        
        await db.invoices.update_one(
            {"id": payment.invoice_id},
            {"$set": {
                "paid_amount": paid_amount,
                "payment_status": payment_status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        return {
            "success": True,
            "message": "Payment recorded successfully",
            "paid_amount": paid_amount,
            "payment_status": payment_status
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Add Follow-up Note
@api_router.post("/recovery/follow-up")
async def add_follow_up(follow_up: FollowUpNote):
    """Add a follow-up note for an invoice"""
    try:
        await db.follow_ups.insert_one(follow_up.dict())
        return {"success": True, "message": "Follow-up note added successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Get Follow-up History
@api_router.get("/recovery/follow-ups/{invoice_id}")
async def get_follow_ups(invoice_id: str):
    """Get all follow-ups for an invoice"""
    try:
        follow_ups = await db.follow_ups.find(
            {"invoice_id": invoice_id},
            {"_id": 0}
        ).sort("created_at", -1).to_list(100)
        return follow_ups
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Get Payment History
@api_router.get("/recovery/payments/{invoice_id}")
async def get_payment_history(invoice_id: str):
    """Get payment history for an invoice"""
    try:
        payments = await db.payment_records.find(
            {"invoice_id": invoice_id},
            {"_id": 0}
        ).sort("payment_date", -1).to_list(100)
        return payments
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Send Payment Reminder via WhatsApp
@api_router.post("/recovery/send-reminder/{invoice_id}")
async def send_payment_reminder(invoice_id: str):
    """Send payment reminder via WhatsApp"""
    try:
        invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        today = datetime.now(timezone.utc)
        due_date = datetime.fromisoformat(invoice['due_date'].replace('Z', '+00:00'))
        days_overdue = (today - due_date).days
        
        outstanding = invoice['total_amount'] - invoice.get('paid_amount', 0)
        
        # Create reminder message
        message = f"""🔔 Payment Reminder - Nectar

Dear {invoice['customer_name']},

This is a friendly reminder regarding:

📄 Invoice: {invoice['invoice_number']}
💰 Amount Due: ₹{outstanding:,.2f}
📅 Due Date: {invoice['due_date'][:10]}
⚠️ Days Overdue: {days_overdue} days

Please process the payment at your earliest convenience.

For any queries, contact us.

Thank you!"""

        # Send via WhatsApp
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "http://localhost:3001/send-message",
                json={
                    "phone_number": invoice['customer_phone'],
                    "message": message
                },
                timeout=30.0
            )
            
            if response.json().get('success'):
                # Record follow-up
                follow_up = FollowUpNote(
                    invoice_id=invoice_id,
                    invoice_number=invoice['invoice_number'],
                    follow_up_date=datetime.now(timezone.utc).isoformat(),
                    contact_method="whatsapp",
                    notes=f"Payment reminder sent via WhatsApp. Days overdue: {days_overdue}",
                    status="contacted",
                    recorded_by="System"
                )
                await db.follow_ups.insert_one(follow_up.dict())
                
                return {"success": True, "message": "Payment reminder sent via WhatsApp"}
            else:
                return {"success": False, "message": "Failed to send WhatsApp reminder"}
                
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Get Customer Outstanding Summary
@api_router.get("/recovery/customer-summary")
async def get_customer_summary():
    """Get outstanding summary by customer"""
    try:
        invoices = await db.invoices.find(
            {
                "status": {"$ne": "cancelled"},
                "payment_status": {"$in": ["unpaid", "partially_paid"]}
            },
            {"_id": 0}
        ).to_list(1000)
        
        customer_summary = {}
        today = datetime.now(timezone.utc)
        
        for invoice in invoices:
            customer_name = invoice['customer_name']
            
            if customer_name not in customer_summary:
                customer_summary[customer_name] = {
                    "customer_name": customer_name,
                    "customer_phone": invoice['customer_phone'],
                    "customer_email": invoice.get('customer_email', ''),
                    "total_outstanding": 0,
                    "invoice_count": 0,
                    "oldest_invoice_days": 0,
                    "invoices": []
                }
            
            outstanding = invoice['total_amount'] - invoice.get('paid_amount', 0)
            due_date = datetime.fromisoformat(invoice['due_date'].replace('Z', '+00:00'))
            days_overdue = (today - due_date).days
            
            if days_overdue > 0:
                customer_summary[customer_name]['total_outstanding'] += outstanding
                customer_summary[customer_name]['invoice_count'] += 1
                customer_summary[customer_name]['oldest_invoice_days'] = max(
                    customer_summary[customer_name]['oldest_invoice_days'],
                    days_overdue
                )
                customer_summary[customer_name]['invoices'].append({
                    "invoice_number": invoice['invoice_number'],
                    "invoice_date": invoice['invoice_date'],
                    "due_date": invoice['due_date'],
                    "days_overdue": days_overdue,
                    "outstanding": outstanding
                })
        
        # Convert to list and sort by total outstanding
        result = list(customer_summary.values())
        result.sort(key=lambda x: x['total_outstanding'], reverse=True)
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
